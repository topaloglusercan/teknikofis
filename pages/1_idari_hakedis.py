"""
İdari Hakediş & Teyit Matrisi - Web Modülü

Düzeltmeler:
- Streamlit data_editor verileri kalıcı session_state içinde tutulur.
- Aynı Excel/JSON dosyası her yeniden çalışmada tekrar yüklenmez.
- Endeks değerleri en az 6 ondalık basamakla, budanmadan gösterilir.
- Eksik endeks aylarında gelecekteki/son satırı kullanmak yerine en yakın önceki ay kullanılır.
- Tarih, kümülatif değer, katsayı ve sütun kontrolleri eklendi.
- Hesaplama sonuçları yeni veri girilene kadar ekranda kalır.
"""

from __future__ import annotations

import datetime
import hashlib
import io
import json
import re
import warnings
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, getcontext
from typing import Any

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

getcontext().prec = 28
warnings.filterwarnings("ignore")

MISSING_TEXTS = {"", "none", "nan", "nat", "<na>"}
ENDKS_MIN_DECIMALS = 6
DEFAULT_ENDKS_MAP = {
    "a": "I o",
    "b1": "Ç o",
    "b2": "D o",
    "b3": "Y o",
    "b4": "K o",
    "b5": "G o",
    "c": "M o",
}

MONTH_NAMES = [
    "",
    "Ocak",
    "Şubat",
    "Mart",
    "Nisan",
    "Mayıs",
    "Haziran",
    "Temmuz",
    "Ağustos",
    "Eylül",
    "Ekim",
    "Kasım",
    "Aralık",
]

MONTH_NUMBERS = {
    "oca": "01",
    "ocak": "01",
    "şub": "02",
    "şubat": "02",
    "sub": "02",
    "subat": "02",
    "mar": "03",
    "mart": "03",
    "nis": "04",
    "nisan": "04",
    "may": "05",
    "mayıs": "05",
    "mayis": "05",
    "haz": "06",
    "haziran": "06",
    "tem": "07",
    "temmuz": "07",
    "ağu": "08",
    "ağustos": "08",
    "agu": "08",
    "agustos": "08",
    "eyl": "09",
    "eylül": "09",
    "eylul": "09",
    "eki": "10",
    "ekim": "10",
    "kas": "11",
    "kasım": "11",
    "kasim": "11",
    "ara": "12",
    "aralık": "12",
    "aralik": "12",
}


# ==========================================
# 1. GÖRSEL TEMİZLİK VE FORMATLAMA
# ==========================================
def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() in MISSING_TEXTS


def clean_decimal(value: Any) -> Decimal:
    """Türk ve uluslararası sayı yazımlarını Decimal'a çevirir."""
    if _is_missing(value):
        return Decimal("0")

    if isinstance(value, Decimal):
        return Decimal("0") if value.is_nan() else value

    if isinstance(value, bool):
        return Decimal(int(value))

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        if pd.isna(value):
            return Decimal("0")
        return Decimal(str(value))

    text = (
        str(value)
        .strip()
        .replace("\u00a0", "")
        .replace(" ", "")
        .replace("TL", "")
        .replace("₺", "")
        .replace("%", "")
    )

    negative_parentheses = text.startswith("(") and text.endswith(")")
    if negative_parentheses:
        text = text[1:-1]

    if "." in text and "," in text:
        # Son görülen ayraç ondalık ayraç kabul edilir.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        # Birden fazla nokta varsa bunlar binlik ayraçtır.
        text = text.replace(".", "")

    try:
        result = Decimal(text)
        if result.is_nan():
            return Decimal("0")
        return -result if negative_parentheses else result
    except (InvalidOperation, ValueError):
        return Decimal("0")


def decimal_to_tr_text(
    value: Any,
    *,
    min_decimals: int = 0,
    max_decimals: int = 15,
    keep_user_text: bool = False,
) -> str:
    """Sayıyı yuvarlamadan Türkçe ondalık gösterime çevirir."""
    if _is_missing(value):
        return ""

    original = str(value).strip()
    if keep_user_text and isinstance(value, str) and "," in original:
        # Kullanıcının zaten Türkçe biçimde yazdığı basamakları aynen koru.
        return original

    number = clean_decimal(value)
    plain = format(number, "f")
    sign = ""
    if plain.startswith("-"):
        sign, plain = "-", plain[1:]

    integer_part, dot, fraction = plain.partition(".")
    if len(fraction) > max_decimals:
        # UI gösteriminde keyfî yuvarlama yapmamak için kesmek yerine Decimal
        # hassasiyet sınırına göre sabit biçim kullanılır. Normal kullanımda bu
        # dal yalnızca çok uzun kayan nokta artıkları için çalışır.
        fraction = fraction[:max_decimals]

    fraction = fraction.rstrip("0")
    if len(fraction) < min_decimals:
        fraction += "0" * (min_decimals - len(fraction))

    if fraction:
        return f"{sign}{integer_part},{fraction}"
    return f"{sign}{integer_part}"


def clean_df_for_ui(df: pd.DataFrame, table_name: str | None = None) -> pd.DataFrame:
    """Excel/JSON verilerini düzenlenebilir metin hücrelerine hazırlar."""
    df_clean = df.copy()
    table_name = (table_name or "").lower()

    for column in df_clean.columns:
        column_upper = str(column).strip().upper()
        is_text_column = column_upper in {"AYLAR", "AĞIRLIK", "ENDEKS SÜTUNU"}
        values: list[str] = []

        for value in df_clean[column]:
            if _is_missing(value):
                values.append("")
                continue

            if isinstance(value, (pd.Timestamp, datetime.datetime, datetime.date)):
                values.append(f"{MONTH_NAMES[value.month]} {str(value.year)[-2:]}")
                continue

            text = str(value).strip()
            if text.endswith(" 00:00:00"):
                try:
                    date_value = pd.to_datetime(text)
                    values.append(f"{MONTH_NAMES[date_value.month]} {str(date_value.year)[-2:]}")
                    continue
                except (TypeError, ValueError):
                    pass

            if is_text_column:
                values.append(text)
            elif table_name == "endeks":
                values.append(
                    decimal_to_tr_text(
                        value,
                        min_decimals=ENDKS_MIN_DECIMALS,
                        max_decimals=15,
                        keep_user_text=True,
                    )
                )
            else:
                values.append(
                    decimal_to_tr_text(
                        value,
                        min_decimals=0,
                        max_decimals=15,
                        keep_user_text=True,
                    )
                )

        df_clean[column] = values

    return df_clean.astype(str)


def get_text_config(df: pd.DataFrame) -> dict[str, Any]:
    return {
        column: st.column_config.TextColumn(
            str(column),
            help="Değer hücreleri metin olarak tutulur; virgüllü sayılar kaybolmaz.",
        )
        for column in df.columns
    }


def tr_format(value: Any, decimals: int = 2) -> str:
    if _is_missing(value):
        return ""
    number = clean_decimal(value)
    formatted = f"{number:,.{decimals}f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def format_pn(value: Any) -> str:
    if _is_missing(value):
        return ""
    number = clean_decimal(value)
    return format(number, ".15f").rstrip("0").rstrip(".").replace(".", ",")


def filter_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    first_column = df.iloc[:, 0].astype(str).str.strip().str.lower()
    return df.loc[~first_column.isin(MISSING_TEXTS)].copy()


# ==========================================
# 2. TARİH VE DOĞRULAMA YARDIMCILARI
# ==========================================
def parse_turkish_date(date_value: Any) -> str | pd.NaT:
    """Oca 22, Ocak 2022, 2022-01 ve Excel tarihlerini YYYY-MM'e çevirir."""
    if _is_missing(date_value):
        return pd.NaT

    if isinstance(date_value, (pd.Timestamp, datetime.datetime, datetime.date)):
        return f"{date_value.year:04d}-{date_value.month:02d}"

    text = str(date_value).strip().lower().replace(".", " ")
    text = re.sub(r"\s+", " ", text)

    iso_match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", text)
    if iso_match:
        year, month = int(iso_match.group(1)), int(iso_match.group(2))
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}"
        return pd.NaT

    month_year_match = re.fullmatch(r"(\d{1,2})[-/](\d{2}|\d{4})", text)
    if month_year_match:
        month = int(month_year_match.group(1))
        year_text = month_year_match.group(2)
        year = int(year_text) if len(year_text) == 4 else 2000 + int(year_text)
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}"
        return pd.NaT

    parts = text.split()
    if len(parts) >= 2 and parts[0] in MONTH_NUMBERS:
        year_text = re.sub(r"\D", "", parts[1])
        if len(year_text) == 2:
            year_text = f"20{year_text}"
        if len(year_text) == 4:
            return f"{year_text}-{MONTH_NUMBERS[parts[0]]}"
        return pd.NaT

    full_date_match = re.fullmatch(r"(\d{1,2})[-/ ](\d{1,2})[-/ ](\d{2}|\d{4})", text)
    if full_date_match:
        day = int(full_date_match.group(1))
        month = int(full_date_match.group(2))
        year_text = full_date_match.group(3)
        year = int(year_text) if len(year_text) == 4 else 2000 + int(year_text)
        try:
            parsed = datetime.date(year, month, day)
            return f"{parsed.year:04d}-{parsed.month:02d}"
        except ValueError:
            return pd.NaT

    return pd.NaT


def _prepare_period_table(
    df: pd.DataFrame,
    *,
    date_column: str,
    table_label: str,
    duplicate_policy: str = "error",
) -> pd.DataFrame:
    prepared = filter_empty_rows(df)
    if prepared.empty:
        return prepared

    if date_column not in prepared.columns:
        raise ValueError(f"{table_label} tablosunda '{date_column}' sütunu bulunamadı.")

    prepared["AyKodu"] = pd.to_datetime(
        prepared[date_column].apply(parse_turkish_date), errors="coerce"
    ).dt.to_period("M")

    invalid_rows = prepared[prepared["AyKodu"].isna()]
    if not invalid_rows.empty:
        row_numbers = ", ".join(str(index + 1) for index in invalid_rows.index[:5])
        raise ValueError(
            f"{table_label} tablosunda geçersiz ay değeri var. Satır(lar): {row_numbers}."
        )

    duplicates = prepared[prepared["AyKodu"].duplicated(keep=False)]
    if not duplicates.empty:
        duplicate_months = ", ".join(sorted({str(value) for value in duplicates["AyKodu"]}))
        if duplicate_policy == "last":
            prepared = prepared.drop_duplicates(subset=["AyKodu"], keep="last")
        else:
            raise ValueError(
                f"{table_label} tablosunda aynı ay birden fazla kez girilmiş: {duplicate_months}."
            )

    return prepared.sort_values("AyKodu").reset_index(drop=True)


def _asof_row(indexed_df: pd.DataFrame, period: pd.Period, table_label: str) -> pd.Series:
    """İstenen ay yoksa, o aydan önceki en yakın satırı döndürür."""
    eligible = indexed_df.index[indexed_df.index <= period]
    if len(eligible) == 0:
        first_month = indexed_df.index.min()
        raise ValueError(
            f"{table_label} tablosu {period} ayını karşılamıyor. "
            f"İlk mevcut ay {first_month}; daha eski endeks/B değeri gerekli."
        )
    return indexed_df.loc[eligible.max()]


def _validate_non_decreasing(df: pd.DataFrame, column: str, label: str) -> None:
    values = [clean_decimal(value) for value in df[column]]
    decreases = [i for i in range(1, len(values)) if values[i] < values[i - 1]]
    if decreases:
        rows = ", ".join(str(i + 1) for i in decreases[:5])
        raise ValueError(
            f"{label} kümülatif olmasına rağmen azalan değer içeriyor. Sorunlu satır(lar): {rows}."
        )


def validate_tables(
    df_prog: pd.DataFrame,
    df_endeks: pd.DataFrame,
    df_alt: pd.DataFrame,
    df_b: pd.DataFrame,
) -> list[str]:
    """Hesabı bozmayacak fakat kullanıcıya gösterilmesi yararlı uyarıları döndürür."""
    warnings_list: list[str] = []

    if not df_alt.empty and "Katsayı" in df_alt.columns:
        coefficient_sum = sum((clean_decimal(value) for value in df_alt["Katsayı"]), Decimal("0"))
        if abs(coefficient_sum - Decimal("1")) > Decimal("0.000001"):
            warnings_list.append(
                f"Alt endeks katsayıları toplamı {decimal_to_tr_text(coefficient_sum)}. "
                "Normalde 1 olmalıdır; hesap yine çalıştırılmayacaktır."
            )

    return warnings_list


# ==========================================
# 3. HESAPLAMA MOTORU
# ==========================================
def hesapla(
    df_prog: pd.DataFrame,
    df_endeks: pd.DataFrame,
    df_alt: pd.DataFrame,
    df_b: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df_prog = filter_empty_rows(df_prog.copy())
    df_endeks = filter_empty_rows(df_endeks.copy())
    df_alt = filter_empty_rows(df_alt.copy())
    df_b = filter_empty_rows(df_b.copy())

    if df_prog.empty or df_endeks.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df_prog.columns = df_prog.columns.astype(str).str.strip()
    df_endeks.columns = df_endeks.columns.astype(str).str.strip()
    df_alt.columns = df_alt.columns.astype(str).str.strip()
    df_b.columns = df_b.columns.astype(str).str.strip()

    required_program_columns = {"AYLAR", "İŞ PROGRAMI KÜMÜLATİF", "İMALAT TUTARI KÜMÜLATİF"}
    missing_program = required_program_columns.difference(df_prog.columns)
    if missing_program:
        raise ValueError(
            "İş Programı tablosunda eksik sütun var: " + ", ".join(sorted(missing_program))
        )

    required_alt_columns = {"Ağırlık", "Katsayı", "Temel Endeks"}
    missing_alt = required_alt_columns.difference(df_alt.columns)
    if missing_alt:
        raise ValueError(
            "Alt Endeks tablosunda eksik sütun var: " + ", ".join(sorted(missing_alt))
        )

    if "AYLAR" not in df_endeks.columns:
        raise ValueError("Endeks tablosunda 'AYLAR' sütunu bulunamadı.")
    if "AYLAR" not in df_b.columns or "B" not in df_b.columns:
        raise ValueError("B tablosunda 'AYLAR' ve 'B' sütunları bulunmalıdır.")

    df_prog = _prepare_period_table(
        df_prog,
        date_column="AYLAR",
        table_label="İş Programı",
    )
    df_endeks = _prepare_period_table(
        df_endeks,
        date_column="AYLAR",
        table_label="Endeks",
    )
    df_b = _prepare_period_table(
        df_b,
        date_column="AYLAR",
        table_label="B",
        duplicate_policy="last",
    )

    if df_prog.empty or df_endeks.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    program_column = "İŞ PROGRAMI KÜMÜLATİF"
    production_column = "İMALAT TUTARI KÜMÜLATİF"
    _validate_non_decreasing(df_prog, program_column, "İş programı")
    _validate_non_decreasing(df_prog, production_column, "İmalat tutarı")

    coefficient_sum = sum((clean_decimal(value) for value in df_alt["Katsayı"]), Decimal("0"))
    if abs(coefficient_sum - Decimal("1")) > Decimal("0.000001"):
        raise ValueError(
            "Alt endeks katsayıları toplamı 1 olmalıdır. "
            f"Mevcut toplam: {decimal_to_tr_text(coefficient_sum)}"
        )

    coefficients = {
        str(row["Ağırlık"]).strip().lower(): clean_decimal(row["Katsayı"])
        for _, row in df_alt.iterrows()
    }
    base_indexes = {
        str(row["Ağırlık"]).strip().lower(): clean_decimal(row["Temel Endeks"])
        for _, row in df_alt.iterrows()
    }

    if "Endeks Sütunu" in df_alt.columns:
        index_map = {
            str(row["Ağırlık"]).strip().lower(): str(row["Endeks Sütunu"]).strip()
            for _, row in df_alt.iterrows()
            if str(row.get("Endeks Sütunu", "")).strip().lower() not in MISSING_TEXTS
        }
        if not index_map:
            index_map = DEFAULT_ENDKS_MAP.copy()
    else:
        index_map = DEFAULT_ENDKS_MAP.copy()

    for weight_key, index_column in index_map.items():
        coefficient = coefficients.get(weight_key, Decimal("0"))
        if coefficient <= 0:
            continue
        if index_column not in df_endeks.columns:
            raise ValueError(
                f"'{weight_key}' ağırlığı için seçilen '{index_column}' sütunu Endeks tablosunda yok."
            )
        if base_indexes.get(weight_key, Decimal("0")) <= 0:
            raise ValueError(
                f"'{weight_key}' ağırlığının Temel Endeks değeri sıfırdan büyük olmalıdır."
            )

    endeks_indexed = df_endeks.set_index("AyKodu").sort_index()
    b_indexed = df_b.set_index("AyKodu").sort_index()
    last_index_month = endeks_indexed.index.max()

    final_program_total = clean_decimal(df_prog.iloc[-1][program_column])
    max_production_total = max(
        (clean_decimal(value) for value in df_prog[production_column]),
        default=Decimal("0"),
    )
    if max_production_total > final_program_total:
        raise ValueError(
            "İmalat kümülatifi, iş programının toplam ödeneğini aşıyor. "
            "Bu durumda tutarın bir bölümü hiçbir ödenek ayına dağıtılamaz."
        )

    buckets: list[dict[str, Any]] = []
    previous_program_total = Decimal("0")
    for _, row in df_prog.iterrows():
        cumulative = clean_decimal(row[program_column])
        capacity = cumulative - previous_program_total
        buckets.append(
            {
                "ay": row["AyKodu"],
                "kapasite": max(capacity, Decimal("0")),
            }
        )
        previous_program_total = cumulative

    cumulative_results: list[float] = []
    detail_rows: list[dict[str, Any]] = []
    previous_production_total = Decimal("0")
    cumulative_price_difference = Decimal("0")

    for _, row in df_prog.iterrows():
        application_month: pd.Period = row["AyKodu"]
        current_production_total = clean_decimal(row[production_column])
        monthly_production = current_production_total - previous_production_total

        if monthly_production <= 0:
            cumulative_results.append(float(cumulative_price_difference))
            previous_production_total = current_production_total
            continue

        effective_application_month = min(application_month, last_index_month)
        application_indexes = _asof_row(
            endeks_indexed,
            effective_application_month,
            "Endeks",
        )

        try:
            b_row = _asof_row(b_indexed, application_month, "B")
            b_factor = clean_decimal(b_row.get("B", 1))
        except ValueError:
            b_factor = Decimal("1")
        if b_factor <= 0:
            b_factor = Decimal("1")

        monthly_price_difference = Decimal("0")
        remaining = monthly_production

        for bucket in buckets:
            if remaining <= 0:
                break
            if bucket["kapasite"] <= 0:
                continue

            used_amount = min(remaining, bucket["kapasite"])
            effective_program_month = min(bucket["ay"], last_index_month)
            delayed = bucket["ay"] < application_month

            if delayed:
                comparison_month = min(effective_application_month, effective_program_month)
                program_indexes = _asof_row(endeks_indexed, comparison_month, "Endeks")
            else:
                program_indexes = application_indexes

            pn = Decimal("0")
            for weight_key, index_column in index_map.items():
                coefficient = coefficients.get(weight_key, Decimal("0"))
                if coefficient == 0:
                    continue

                base_index = base_indexes.get(weight_key, Decimal("0"))
                application_index = clean_decimal(application_indexes.get(index_column, 0))
                program_index = clean_decimal(program_indexes.get(index_column, 0))
                valid_index = min(application_index, program_index) if delayed else application_index

                if base_index > 0:
                    pn += coefficient * (valid_index / base_index)
                else:
                    pn += coefficient

            price_difference_slice = (
                used_amount * b_factor * (pn - Decimal("1"))
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            detail_rows.append(
                {
                    "Hakediş Ayı": str(application_month),
                    "İş Programı (Ödenek) Ayı": str(bucket["ay"]),
                    "Kullanılan Tutar": float(used_amount),
                    "Uygulanan Pn (15 Hane)": pn,
                    "Fiyat Farkı Tutarı": float(price_difference_slice),
                }
            )

            monthly_price_difference += price_difference_slice
            bucket["kapasite"] -= used_amount
            remaining -= used_amount

        if remaining > 0:
            raise RuntimeError(
                f"{application_month} ayında {tr_format(remaining)} TL tutar ödenek dilimine dağıtılamadı."
            )

        cumulative_price_difference += monthly_price_difference
        cumulative_results.append(float(cumulative_price_difference))
        previous_production_total = current_production_total

    result_df = df_prog.drop(columns=["AyKodu"]).copy()
    result_df["KÜMÜLATİF FİYAT FARKI"] = cumulative_results
    detail_df = pd.DataFrame(detail_rows)

    if not detail_df.empty:
        pivot_df = detail_df.pivot_table(
            index="Hakediş Ayı",
            columns="İş Programı (Ödenek) Ayı",
            values="Kullanılan Tutar",
            aggfunc="sum",
            fill_value=0,
        )
        pivot_df["HAKEDİŞ TUTARI (Toplam)"] = pivot_df.sum(axis=1)
        pivot_df.loc["ÖDENEK MİKTARI"] = pivot_df.sum()
    else:
        pivot_df = pd.DataFrame()

    return result_df, pivot_df, detail_df


# ==========================================
# 4. EXCEL ŞABLON MOTORU
# ==========================================
def _excel_value(value: Any, *, force_text: bool = False) -> Any:
    if _is_missing(value):
        return ""
    if force_text:
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime.datetime, datetime.date)):
        return value
    number = clean_decimal(value)
    if number != 0 or str(value).strip() in {"0", "0,0", "0,00", "0.0", "0,000000"}:
        return float(number)
    return str(value)


def generate_excel_download(
    df_prog: pd.DataFrame,
    df_endeks: pd.DataFrame,
    df_alt: pd.DataFrame,
    df_b: pd.DataFrame,
) -> bytes:
    """Ekrandaki tablolarla yeniden yüklenebilir Excel şablonu üretir."""
    workbook = Workbook()
    thin = Side(style="thin", color="B0C4DE")

    def border() -> Border:
        return Border(top=thin, left=thin, right=thin, bottom=thin)

    def fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def font(color: str = "1A1A2E", bold: bool = False, size: int = 10) -> Font:
        return Font(color=color, bold=bold, size=size, name="Calibri")

    def alignment(horizontal: str = "left", wrap: bool = False) -> Alignment:
        return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)

    header_color = "2C5F8A"
    white = "FFFFFF"
    yellow = "FFFDE7"
    note_color = "FFF9C4"

    def make_header(sheet: Any, title: str, note: str = "") -> None:
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:Z1")
        sheet["A1"] = title
        sheet["A1"].fill = fill(header_color)
        sheet["A1"].font = font(white, True, 12)
        sheet["A1"].alignment = alignment("center")
        sheet.row_dimensions[1].height = 22
        if note:
            sheet.merge_cells("A2:Z2")
            sheet["A2"] = note
            sheet["A2"].fill = fill(note_color)
            sheet["A2"].font = font("7B5800", False, 9)
            sheet["A2"].alignment = alignment("center")
            sheet.row_dimensions[2].height = 14

    def column_header(sheet: Any, row: int, headers: list[str], widths: list[int]) -> None:
        for column_index, (header, width) in enumerate(zip(headers, widths), 1):
            cell = sheet.cell(row, column_index, header)
            cell.fill = fill(header_color)
            cell.font = font(white, True, 9)
            cell.alignment = alignment("center")
            cell.border = border()
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.row_dimensions[row].height = 16

    def write_df(
        sheet: Any,
        df: pd.DataFrame,
        start_row: int,
        yellow_columns: list[int],
        table_name: str,
    ) -> None:
        for row_index, row in enumerate(df.itertuples(index=False, name=None), start_row):
            alternating = yellow if row_index % 2 == 0 else "FFFEF0"
            for column_index, value in enumerate(row, 1):
                force_text = str(df.columns[column_index - 1]).strip().upper() in {
                    "AYLAR", "AĞIRLIK", "ENDEKS SÜTUNU"
                }
                cell = sheet.cell(
                    row_index,
                    column_index,
                    _excel_value(value, force_text=force_text),
                )
                cell.fill = fill(yellow if column_index in yellow_columns else alternating)
                cell.font = font("1A1A2E", False, 9)
                cell.alignment = alignment("right" if column_index > 1 else "left")
                cell.border = border()
                if column_index > 1 and isinstance(cell.value, (int, float)):
                    if table_name == "endeks":
                        cell.number_format = "#,##0.000000############"
                    elif table_name == "prog":
                        cell.number_format = "#,##0.00############"
                    else:
                        cell.number_format = "0.###############"

        for row_index in range(start_row + len(df), start_row + len(df) + 10):
            alternating = yellow if row_index % 2 == 0 else "FFFEF0"
            for column_index in range(1, len(df.columns) + 1):
                cell = sheet.cell(row_index, column_index, "")
                cell.fill = fill(yellow if column_index in yellow_columns else alternating)
                cell.font = font("1A1A2E", False, 9)
                cell.alignment = alignment("right" if column_index > 1 else "left")
                cell.border = border()

    sheet_program = workbook.active
    sheet_program.title = "IsProgrami"
    make_header(
        sheet_program,
        "1. İŞ PROGRAMI VE İMALATLAR",
        "Sarı hücrelere verileri girin. Ay formatı: Oca 22, Şub 22, ... / Tüm tutarlar TL",
    )
    column_header(sheet_program, 3, list(df_prog.columns), [16, 30, 30])
    write_df(sheet_program, df_prog, 4, [1, 2, 3], "prog")

    sheet_index = workbook.create_sheet("Endeks")
    make_header(
        sheet_index,
        "2. ENDEKS TABLOSU",
        "Kaynak: TÜİK / Çevre ve Şehircilik Bakanlığı yayınları",
    )
    column_header(
        sheet_index,
        3,
        list(df_endeks.columns),
        [16] + [15] * (len(df_endeks.columns) - 1),
    )
    write_df(
        sheet_index,
        df_endeks,
        4,
        list(range(1, len(df_endeks.columns) + 1)),
        "endeks",
    )

    sheet_alt = workbook.create_sheet("AltEndeks")
    make_header(
        sheet_alt,
        "3. ALT ENDEKS AĞIRLIKLARI",
        "Katsayılar toplamı = 1,00 olmalıdır. Temel Endeks = sözleşme tarihindeki endeks değeri.",
    )
    column_header(
        sheet_alt,
        3,
        list(df_alt.columns),
        [12, 14, 16, 16][: len(df_alt.columns)],
    )
    write_df(
        sheet_alt,
        df_alt,
        4,
        list(range(1, len(df_alt.columns) + 1)),
        "alt",
    )

    sheet_b = workbook.create_sheet("B")
    make_header(
        sheet_b,
        "4. B KATSAYISI TABLOSU",
        "Müteahhit fiyat farkı katsayısı (genellikle 1,00).",
    )
    column_header(sheet_b, 3, list(df_b.columns), [16, 14])
    write_df(sheet_b, df_b, 4, [1, 2], "b")

    guide_sheet = workbook.create_sheet("Kilavuz")
    guide_sheet.sheet_view.showGridLines = False
    guide_sheet.column_dimensions["A"].width = 90
    guide_sheet["A1"] = "İDARİ HAKEDİŞ ŞABLONU — KULLANIM KILAVUZU"
    guide_sheet["A1"].fill = fill(header_color)
    guide_sheet["A1"].font = font(white, True, 14)
    guide_sheet["A1"].alignment = alignment("center")
    guide_sheet.row_dimensions[1].height = 28

    help_text = [
        "",
        "💡 EXCEL ŞABLONU İLE VERİ GİRİŞİ",
        "1. Bu Excel dosyasındaki sarı alanları kendi projenizin değerleriyle doldurun.",
        "2. Programın sol menüsündeki 'Excel Dosyası Seç (.xlsx)' bölümünden bu dosyayı yükleyin.",
        "3. Veriler tabloya aktarılır. Ardından 'Hesapla' butonuna basabilirsiniz.",
        "4. Endeks hücreleri en az 6 ondalık basamak gösterecek şekilde ayarlanmıştır.",
    ]
    for row_index, text in enumerate(help_text, 2):
        cell = guide_sheet.cell(row_index, 1, text)
        if text.startswith("💡") or text.startswith("⚠️"):
            cell.font = font("1A3A6E", True, 11)
            cell.fill = fill("D6E4F0")
        else:
            cell.font = font("1A1A2E", False, 10)
        guide_sheet.row_dimensions[row_index].height = 18

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _find_header_row(raw_df: pd.DataFrame) -> int:
    for row_index, row in raw_df.iterrows():
        normalized = {str(value).strip().upper() for value in row.values if not _is_missing(value)}
        if "AYLAR" in normalized or "AĞIRLIK" in normalized:
            return int(row_index)
    return 0


def load_from_excel(file: Any) -> dict[str, pd.DataFrame]:
    excel = pd.ExcelFile(file)
    dataframes: dict[str, pd.DataFrame] = {}
    sheet_map = {
        "IsProgrami": ("prog_df", "prog"),
        "Endeks": ("endeks_df", "endeks"),
        "AltEndeks": ("alt_df", "alt"),
        "B": ("b_df", "b"),
    }

    for sheet_name in excel.sheet_names:
        if sheet_name not in sheet_map:
            continue

        raw_preview = pd.read_excel(excel, sheet_name=sheet_name, header=None, nrows=12)
        header_row = _find_header_row(raw_preview)
        df = pd.read_excel(excel, sheet_name=sheet_name, header=header_row)
        df = df.loc[:, ~df.columns.astype(str).str.contains(r"^Unnamed", regex=True)]
        df = df.dropna(how="all")

        if sheet_name == "AltEndeks" and "Endeks Sütunu" not in df.columns:
            df["Endeks Sütunu"] = (
                df["Ağırlık"]
                .astype(str)
                .str.strip()
                .str.lower()
                .map(DEFAULT_ENDKS_MAP)
                .fillna("")
            )

        state_key, table_name = sheet_map[sheet_name]
        dataframes[state_key] = clean_df_for_ui(df, table_name)

    return dataframes


# ==========================================
# 5. STATE VE ARAYÜZ YARDIMCILARI
# ==========================================
def _default_tables() -> dict[str, pd.DataFrame]:
    return {
        "prog_df": pd.DataFrame(
            {
                "AYLAR": ["Oca 22"],
                "İŞ PROGRAMI KÜMÜLATİF": ["0,00"],
                "İMALAT TUTARI KÜMÜLATİF": ["0,00"],
            }
        ),
        "endeks_df": pd.DataFrame(
            {
                "AYLAR": ["Oca 22"],
                "I o": ["0,000000"],
                "Ç o": ["0,000000"],
                "D o": ["0,000000"],
                "Y o": ["0,000000"],
                "K o": ["0,000000"],
                "G o": ["0,000000"],
                "M o": ["0,000000"],
            }
        ),
        "alt_df": pd.DataFrame(
            {
                "Ağırlık": ["a", "b1", "b2", "b3", "b4", "b5", "c"],
                "Katsayı": ["0"] * 7,
                "Temel Endeks": ["0"] * 7,
                "Endeks Sütunu": ["I o", "Ç o", "D o", "Y o", "K o", "G o", "M o"],
            }
        ),
        "b_df": pd.DataFrame({"AYLAR": ["Oca 22"], "B": ["1"]}),
    }


def initialize_state() -> None:
    defaults = _default_tables()
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    state_defaults = {
        "editor_version": 0,
        "loaded_excel_hash": None,
        "loaded_json_hash": None,
        "calculation_result": None,
        "calculation_signature": None,
        "calculation_error": None,
    }
    for key, value in state_defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def dataframe_signature(*dataframes: pd.DataFrame) -> str:
    payload = [
        df.fillna("").astype(str).to_dict(orient="split")
        for df in dataframes
    ]
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _replace_tables(dataframes: dict[str, pd.DataFrame]) -> None:
    for key in ("prog_df", "endeks_df", "alt_df", "b_df"):
        if key in dataframes:
            st.session_state[key] = dataframes[key].copy()
    st.session_state.editor_version += 1
    st.session_state.calculation_result = None
    st.session_state.calculation_signature = None
    st.session_state.calculation_error = None


def _file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _sync_editor_result(state_key: str, edited_df: pd.DataFrame) -> None:
    st.session_state[state_key] = edited_df.copy()


def _display_results(
    result_df: pd.DataFrame,
    pivot_df: pd.DataFrame,
    detail_df: pd.DataFrame,
) -> None:
    st.success("✅ Hesaplama başarılı.")
    tab_detail, tab_pivot, tab_result, tab_chart = st.tabs(
        ["🔍 Dilim Detay", "📊 Teyit Matrisi", "📑 Kümülatif Sonuç", "📈 Analiz Grafiği"]
    )

    with tab_detail:
        if detail_df.empty:
            st.info("Hesaplanacak pozitif imalat dilimi bulunamadı.")
        else:
            detail_display = detail_df.copy()
            detail_display["Kullanılan Tutar"] = detail_display["Kullanılan Tutar"].apply(tr_format)
            detail_display["Fiyat Farkı Tutarı"] = detail_display["Fiyat Farkı Tutarı"].apply(tr_format)
            detail_display["Uygulanan Pn (15 Hane)"] = detail_display[
                "Uygulanan Pn (15 Hane)"
            ].apply(format_pn)
            st.dataframe(detail_display, use_container_width=True)

    with tab_pivot:
        if pivot_df.empty:
            st.info("Teyit matrisi için dağıtılmış tutar bulunamadı.")
        else:
            pivot_display = pivot_df.applymap(tr_format)
            styled = pivot_display.style.set_properties(
                subset=["HAKEDİŞ TUTARI (Toplam)"],
                **{"font-weight": "bold", "background-color": "#e6f2ff"},
            )
            st.dataframe(styled, use_container_width=True)

    with tab_result:
        result_display = result_df.copy()
        for column in result_display.columns:
            if any(text in str(column).upper() for text in ("TUTAR", "PROGRAM", "FARKI")):
                result_display[column] = result_display[column].apply(tr_format)
        st.dataframe(result_display, use_container_width=True)

    with tab_chart:
        try:
            program_raw = _prepare_period_table(
                st.session_state.prog_df.copy(),
                date_column="AYLAR",
                table_label="İş Programı",
            )
            months = [str(value) for value in program_raw["AyKodu"]]
            program_values = [
                float(clean_decimal(value))
                for value in program_raw["İŞ PROGRAMI KÜMÜLATİF"]
            ]
            production_values = [
                float(clean_decimal(value))
                for value in program_raw["İMALAT TUTARI KÜMÜLATİF"]
            ]
            price_difference_values = [
                float(clean_decimal(value))
                for value in result_df["KÜMÜLATİF FİYAT FARKI"]
            ]

            x_values = list(range(len(months)))
            figure, (axis_main, axis_difference) = plt.subplots(
                2,
                1,
                figsize=(12, 7),
                sharex=True,
                gridspec_kw={"height_ratios": [3, 1], "hspace": 0.1},
            )
            figure.patch.set_facecolor("#0d1117")

            for axis in (axis_main, axis_difference):
                axis.set_facecolor("#161b22")
                axis.tick_params(colors="#8b949e", labelsize=9)
                for spine in axis.spines.values():
                    spine.set_edgecolor("#30363d")
                axis.grid(color="#21262d", linewidth=0.7, linestyle="--")

            axis_main.plot(
                x_values,
                program_values,
                color="#58a6ff",
                linewidth=2.5,
                marker="o",
                markersize=5,
                label="İş Programı (Kümülatif)",
            )
            axis_main.plot(
                x_values,
                production_values,
                color="#3fb950",
                linewidth=2.5,
                marker="s",
                markersize=5,
                linestyle="--",
                label="İmalat Tutarı (Kümülatif)",
            )
            axis_main.fill_between(
                x_values,
                program_values,
                production_values,
                where=[p >= i for p, i in zip(program_values, production_values)],
                alpha=0.15,
                color="#f85149",
                label="Gecikme",
            )
            axis_main.fill_between(
                x_values,
                program_values,
                production_values,
                where=[p < i for p, i in zip(program_values, production_values)],
                alpha=0.15,
                color="#3fb950",
                label="Öne Geçme",
            )
            axis_main.set_ylabel("Tutar (TL)", color="#8b949e", fontsize=10)
            axis_main.yaxis.set_major_formatter(
                mticker.FuncFormatter(lambda value, _: tr_format(value))
            )
            axis_main.legend(
                fontsize=9,
                facecolor="#21262d",
                edgecolor="#30363d",
                labelcolor="#c9d1d9",
                loc="upper left",
            )
            axis_main.set_title(
                "Kümülatif İş Programı & İmalat Takibi",
                color="#e6edf3",
                fontsize=13,
                pad=12,
            )

            bar_colors = ["#3fb950" if value >= 0 else "#f85149" for value in price_difference_values]
            axis_difference.bar(
                x_values,
                price_difference_values,
                color=bar_colors,
                alpha=0.85,
                width=0.6,
            )
            axis_difference.axhline(0, color="#30363d", linewidth=1)
            axis_difference.set_ylabel("Fiyat Farkı (TL)", color="#8b949e", fontsize=10)
            axis_difference.yaxis.set_major_formatter(
                mticker.FuncFormatter(lambda value, _: tr_format(value))
            )
            axis_difference.set_title(
                "Kümülatif Fiyat Farkı",
                color="#c9d1d9",
                fontsize=11,
                pad=6,
            )
            axis_difference.set_xticks(x_values)
            axis_difference.set_xticklabels(
                months,
                rotation=45,
                ha="right",
                fontsize=9,
                color="#8b949e",
            )

            st.pyplot(figure)
            plt.close(figure)
        except Exception as exc:
            st.error(f"Grafik oluşturulamadı: {exc}")


# ==========================================
# 6. STREAMLIT UYGULAMASI
# ==========================================
def main() -> None:
    st.set_page_config(
        page_title="İdari Hakediş Modülü",
        layout="wide",
        page_icon="📂",
    )
    initialize_state()

    st.title("📂 İdari Hakediş & Teyit Matrisi")

    st.sidebar.markdown("### 📥 Excel ile Proje Yükle")
    uploaded_excel = st.sidebar.file_uploader(
        "Excel Dosyası Seç (.xlsx)",
        type=["xlsx"],
        key="excel_uploader",
    )
    if uploaded_excel is not None:
        excel_bytes = uploaded_excel.getvalue()
        current_hash = _file_hash(excel_bytes)
        if current_hash != st.session_state.loaded_excel_hash:
            try:
                loaded = load_from_excel(io.BytesIO(excel_bytes))
                _replace_tables(loaded)
                st.session_state.loaded_excel_hash = current_hash
                st.sidebar.success("✅ Veriler bir kez yüklendi ve formatlandı.")
                st.rerun()
            except Exception as exc:
                st.sidebar.error(f"Excel yükleme hatası: {exc}")

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📥 Önceki Projeyi Yükle (.json)")
    uploaded_json = st.sidebar.file_uploader(
        "JSON Dosyası Seç",
        type=["json"],
        key="json_uploader",
    )
    if uploaded_json is not None:
        json_bytes = uploaded_json.getvalue()
        current_hash = _file_hash(json_bytes)
        if current_hash != st.session_state.loaded_json_hash:
            try:
                data = json.loads(json_bytes.decode("utf-8-sig"))
                loaded_json: dict[str, pd.DataFrame] = {}
                if "prog" in data:
                    loaded_json["prog_df"] = clean_df_for_ui(pd.DataFrame(data["prog"]), "prog")
                if "endeks" in data:
                    loaded_json["endeks_df"] = clean_df_for_ui(
                        pd.DataFrame(data["endeks"]), "endeks"
                    )
                if "alt" in data:
                    loaded_json["alt_df"] = clean_df_for_ui(pd.DataFrame(data["alt"]), "alt")
                if "b" in data:
                    loaded_json["b_df"] = clean_df_for_ui(pd.DataFrame(data["b"]), "b")
                _replace_tables(loaded_json)
                st.session_state.loaded_json_hash = current_hash
                st.sidebar.success("✅ Proje bir kez yüklendi.")
                st.rerun()
            except Exception as exc:
                st.sidebar.error(f"JSON yükleme hatası: {exc}")

    if st.sidebar.button("🧹 Yeni boş proje", use_container_width=True):
        _replace_tables(_default_tables())
        st.session_state.loaded_excel_hash = None
        st.session_state.loaded_json_hash = None
        st.rerun()

    version = st.session_state.editor_version
    left_column, right_column = st.columns(2)

    with left_column:
        st.subheader("1. İş Programı ve İmalatlar")
        edited_program = st.data_editor(
            st.session_state.prog_df,
            column_config=get_text_config(st.session_state.prog_df),
            num_rows="dynamic",
            use_container_width=True,
            key=f"prog_editor_{version}",
        )
        _sync_editor_result("prog_df", edited_program)

        st.subheader("3. Alt Endeks Ağırlıkları")
        edited_alt = st.data_editor(
            st.session_state.alt_df,
            column_config=get_text_config(st.session_state.alt_df),
            num_rows="dynamic",
            use_container_width=True,
            key=f"alt_editor_{version}",
        )
        _sync_editor_result("alt_df", edited_alt)

    with right_column:
        st.subheader("2. Endeks Tablosu")
        st.caption("Endeks hücreleri en az 6 ondalık basamakla saklanır; iki haneye yuvarlanmaz.")
        edited_index = st.data_editor(
            st.session_state.endeks_df,
            column_config=get_text_config(st.session_state.endeks_df),
            num_rows="dynamic",
            use_container_width=True,
            key=f"endeks_editor_{version}",
        )
        _sync_editor_result("endeks_df", edited_index)

        st.subheader("4. B Katsayısı Tablosu")
        edited_b = st.data_editor(
            st.session_state.b_df,
            column_config=get_text_config(st.session_state.b_df),
            num_rows="dynamic",
            use_container_width=True,
            key=f"b_editor_{version}",
        )
        _sync_editor_result("b_df", edited_b)

    current_signature = dataframe_signature(
        edited_program,
        edited_index,
        edited_alt,
        edited_b,
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📤 Projeyi Dışa Aktar / Şablon Al")
    json_payload = json.dumps(
        {
            "prog": edited_program.to_dict(orient="records"),
            "endeks": edited_index.to_dict(orient="records"),
            "alt": edited_alt.to_dict(orient="records"),
            "b": edited_b.to_dict(orient="records"),
        },
        ensure_ascii=False,
        indent=4,
        default=str,
    )

    json_column, excel_column = st.sidebar.columns(2)
    with json_column:
        st.download_button(
            label="💾 JSON",
            data=json_payload,
            file_name="hakedis_projem.json",
            mime="application/json",
            use_container_width=True,
        )
    with excel_column:
        st.download_button(
            label="📊 EXCEL",
            data=generate_excel_download(
                edited_program,
                edited_index,
                edited_alt,
                edited_b,
            ),
            file_name="idari_hakedis_sablonu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")

    if st.button(
        "🚀 Hesapla ve Sonuçları Göster",
        use_container_width=True,
        type="primary",
    ):
        with st.spinner("Matematiksel motor çalışıyor..."):
            try:
                result = hesapla(
                    edited_program,
                    edited_index,
                    edited_alt,
                    edited_b,
                )
                st.session_state.calculation_result = result
                st.session_state.calculation_signature = current_signature
                st.session_state.calculation_error = None
            except Exception as exc:
                st.session_state.calculation_result = None
                st.session_state.calculation_signature = None
                st.session_state.calculation_error = str(exc)

    if st.session_state.calculation_error:
        st.error(f"Hesaplama hatası: {st.session_state.calculation_error}")

    if st.session_state.calculation_result is not None:
        if st.session_state.calculation_signature != current_signature:
            st.warning(
                "Tablolarda hesaplamadan sonra değişiklik yapıldı. Sonuçlar eski veriye ait; "
                "yeniden 'Hesapla' düğmesine basın."
            )
        else:
            result_df, pivot_df, detail_df = st.session_state.calculation_result
            if result_df.empty:
                st.warning(
                    "Lütfen tablolara geçerli veri girin. İş Programı ve Endeks boş olamaz."
                )
            else:
                _display_results(result_df, pivot_df, detail_df)


if __name__ == "__main__":
    main()

