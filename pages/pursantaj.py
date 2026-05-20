import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import json, copy, io
from datetime import datetime

st.set_page_config(page_title="Pursantaj Yönetim", layout="wide",
                   initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
* { font-family:'Inter',sans-serif; }
[data-testid="stAppViewContainer"] { background:#0d1117; }
[data-testid="stSidebar"] { display:none; }
.block-container { padding:1rem 2.5rem 2rem; max-width:1100px; margin:auto; }
h1 { text-align:center; font-size:1.9rem; font-weight:800; color:#e6edf3; margin:.2rem 0 .8rem; }
.stTabs [data-baseweb="tab-list"] { background:#161b22; border-radius:8px; padding:3px; gap:4px; border:1px solid #30363d; }
.stTabs [data-baseweb="tab"] { color:#8b949e; border-radius:6px; padding:.35rem 1.1rem; font-size:.83rem; font-weight:500; }
.stTabs [aria-selected="true"] { background:#21262d !important; color:#e6edf3 !important; }
.stTabs [data-baseweb="tab-panel"] { padding-top:.8rem; }
div[data-testid="stTextInput"] label, div[data-testid="stNumberInput"] label,
div[data-testid="stSelectbox"] label, div[data-testid="stMultiSelect"] label
  { color:#8b949e !important; font-size:.75rem !important; }
div[data-testid="stTextInput"] input, div[data-testid="stNumberInput"] input {
  background:#161b22 !important; color:#e6edf3 !important;
  border:1px solid #30363d !important; border-radius:7px !important; font-size:.9rem !important; }
.stButton>button { background:#21262d; border:1px solid #30363d; color:#c9d1d9;
  border-radius:7px; font-size:.82rem; font-weight:500; padding:.4rem .7rem;
  width:100%; transition:all .15s; }
.stButton>button:hover { background:#30363d; border-color:#388bfd; color:#e6edf3; }
.stButton>button[kind="primary"] { background:#1a7f37; border-color:#2ea043; color:#fff; }
div[data-testid="stDownloadButton"] button { background:#21262d; border:1px solid #30363d;
  color:#c9d1d9; border-radius:7px; font-size:.82rem; padding:.4rem .7rem; width:100%; }
.mrow { display:flex; gap:.6rem; margin-bottom:.6rem; }
.mbox { flex:1; background:#161b22; border:1px solid #30363d; border-radius:8px; padding:.6rem .9rem; }
.mbox .lbl { font-size:.72rem; color:#8b949e; font-weight:500; }
.mbox .val { font-size:1.05rem; font-weight:700; color:#58a6ff; margin-top:2px; }
.mbox .sub { font-size:.72rem; color:#6e7681; }
.copy-box { background:#161b22; border:1px solid #d29922; border-radius:8px; padding:1rem; margin:.5rem 0; }
.detect-card { background:#161b22; border:1px solid #30363d; border-radius:10px; padding:1rem 1.2rem; margin:.4rem 0; }
.detect-card.col-card { border-left:4px solid #388bfd; }
.detect-card.hier-card { border-left:4px solid #3fb950; }
.detect-card.weight-card { border-left:4px solid #d29922; }
.detect-title { font-size:.8rem; font-weight:700; color:#8b949e; text-transform:uppercase; letter-spacing:.5px; margin-bottom:.5rem; }
.col-row { display:flex; justify-content:space-between; align-items:center; padding:.25rem 0; border-bottom:1px solid #21262d; }
.col-name { font-weight:600; color:#79c0ff; font-size:.85rem; }
.col-desc { color:#c9d1d9; font-size:.85rem; }
</style>
""", unsafe_allow_html=True)

# ── Veri modeli ───────────────────────────────────────────────────────────
def mk(nid,name,parent,pct):
    return {"id":nid,"name":name,"parent":parent,"pct":float(pct)}

def init_tree():
    return [mk(0,"Proje Geneli",None,100.0),
            mk(1,"Üst Yapı",0,70.0),mk(2,"Alt Yapı",0,30.0)]

DEFS = {
    "tree": init_tree(), "sel": 0, "bedel": 1_000_000.0, "next_id": 3,
    "activities": {},   # {node_id: [row, ...]}
    "copy_mode": False,
    "parsed_rows": None, "overrides": {},
    "grup_headers": ["Grup 1", "Grup 2", "Grup 3"],
    "show_grup_headers": False,
}
for k,v in DEFS.items():
    if k not in st.session_state: st.session_state[k] = v

# ── Ağaç yardımcıları ─────────────────────────────────────────────────────
def get_node(nid):
    return next((n for n in st.session_state.tree if n["id"]==nid), None)
def children_of(nid):
    return [n for n in st.session_state.tree if n["parent"]==nid]
def sibling_total_excl(nid):
    n = get_node(nid)
    if not n or n["parent"] is None: return 0.0
    return sum(x["pct"] for x in st.session_state.tree
               if x["parent"]==n["parent"] and x["id"]!=nid)
def global_pct(nid):
    n = get_node(nid)
    if not n: return 0.0
    if n["parent"] is None: return n["pct"]/100.0
    sibs = [x for x in st.session_state.tree if x["parent"]==n["parent"]]
    st_ = sum(s["pct"] for s in sibs)
    return global_pct(n["parent"]) * (n["pct"]/st_ if st_ else 0)
def node_path(nid, sep=" › "):
    """Kökten bu node'a tam yol: 'çşm › k1 › üst › blok'"""
    parts = []
    current = nid
    while current is not None:
        n = get_node(current)
        if not n: break
        parts.append(n["name"])
        current = n["parent"]
    return sep.join(reversed(parts))

def node_status(nid):
    kids = children_of(nid)
    if not kids: return "ok"
    t = sum(k["pct"] for k in kids)
    return "ok" if abs(t-100)<0.05 else ("over" if t>100 else "under")
def is_leaf(nid): return len(children_of(nid))==0
def auto_distribute(pid):
    """
    Largest Remainder Method ile tam olarak %100 dağıtım.
    Hiçbir birikimli yuvarlama hatası kalmaz (5 kopya → hâlâ %100.0000).
    """
    kids = children_of(pid)
    if not kids: return
    n       = len(kids)
    PREC    = 4          # ondalık basamak
    SCALE   = 10**PREC   # 10000
    total   = 100 * SCALE   # 1_000_000
    base    = total // n
    rem     = total - base * n
    # Kalan payları büyükten küçüğe dağıt (tümü eşit ağırlıklı, ama
    # remainder sadece ilk `rem` çocuğa +1 birim gider)
    for i, k in enumerate(kids):
        raw = base + (1 if i < rem else 0)
        k["pct"] = raw / SCALE
def add_child(pid):
    nid = st.session_state.next_id
    st.session_state.tree.append(mk(nid, f"Yeni Öğe {nid}", pid, 0.0))
    st.session_state.next_id += 1; auto_distribute(pid); return nid
def add_sibling(sid):
    n = get_node(sid); pid = n["parent"] if n else None
    nid = st.session_state.next_id
    st.session_state.tree.append(mk(nid, f"Yeni Öğe {nid}", pid, 0.0))
    st.session_state.next_id += 1
    if pid is not None: auto_distribute(pid)
    return nid
def delete_subtree(nid):
    for k in children_of(nid): delete_subtree(k["id"])
    st.session_state.tree[:] = [n for n in st.session_state.tree if n["id"]!=nid]
    st.session_state.activities.pop(nid, None)
def copy_subtree(src_id, new_parent, new_name=None):
    src = get_node(src_id)
    if not src: return None
    src_kids = list(children_of(src_id))
    nid = st.session_state.next_id; st.session_state.next_id += 1
    st.session_state.tree.append(mk(nid, new_name or src["name"], new_parent, src["pct"]))
    if src_id in st.session_state.activities:
        st.session_state.activities[nid] = copy.deepcopy(st.session_state.activities[src_id])
    for ch in src_kids: copy_subtree(ch["id"], nid)
    return nid

# ── Para hesabı ───────────────────────────────────────────────────────────
def compute_act_budgets(act_rows, node_budget):
    """
    Parasal hesap — WBS öğesinin kendi bütçesi baz alınır:
      d4 varsa  → budget = d4 × node_budget   (genel oran × öğe bütçesi)
      Sadece d3 → budget = d3 × parent_budget  (alt oran × ebeveyn parası)
    Sonuçlar kuruş hassasiyetinde (2 ondalık) yuvarlanır — birikimli float hatası önlenir.
    """
    # node_budget'ı da kuruş hassasiyetine yuvarla
    node_budget = round(node_budget, 2)
    no_map = {r["no"]: r for r in act_rows}
    cache  = {}
    def get_bgt(no):
        if no in cache: return cache[no]
        r = no_map.get(no)
        if not r: cache[no]=0.0; return 0.0
        if r.get("d4") is not None:
            b = round(r["d4"] * node_budget, 2)
        elif r.get("d3") is not None:
            parts = no.split(".")
            parent_no = ".".join(parts[:-1])
            parent_bgt = get_bgt(parent_no) if parent_no else node_budget
            b = round(r["d3"] * parent_bgt, 2)
        else:
            b = 0.0
        cache[no] = b; return b
    return {r["no"]: get_bgt(r["no"]) for r in act_rows}

# ── JSON kaydet / yükle ───────────────────────────────────────────────────
def export_json():
    acts_serializable = {str(k): v for k,v in st.session_state.activities.items()}
    return json.dumps({
        "tree": [{"id":n["id"],"name":n["name"],"parent":n["parent"],"pct":n["pct"]}
                 for n in st.session_state.tree],
        "activities": acts_serializable,
        "bedel": st.session_state.bedel,
        "grup_headers": st.session_state.grup_headers,
        "cons_decisions": st.session_state.get("cons_decisions", {}),
        "cons_rename_map": st.session_state.get("cons_rename_map", {}),
        "cons_confirmed": st.session_state.get("cons_confirmed", False),
    }, ensure_ascii=False, indent=2).encode("utf-8")

def import_json(data):
    """Eski format (düz liste) ve yeni format (dict) ikisini de destekler."""
    # Eski format: JSON doğrudan [{"id":0,"name":"...",...}] listesi
    if isinstance(data, list):
        tree_data    = data
        raw_acts     = {}
        bedel_val    = 1_000_000.0
        gh_val       = ["Grup 1","Grup 2","Grup 3"]
    else:
        tree_data    = data.get("tree", [])
        raw_acts     = data.get("activities", {})
        bedel_val    = data.get("bedel", 1_000_000.0)
        gh_val       = data.get("grup_headers", ["Grup 1","Grup 2","Grup 3"])

    st.session_state.tree          = [mk(n["id"],n["name"],n["parent"],n["pct"]) for n in tree_data]
    st.session_state.next_id       = max(n["id"] for n in tree_data) + 1
    st.session_state.sel           = next((n["id"] for n in tree_data if n["parent"] is None), 0)
    st.session_state.bedel         = bedel_val
    st.session_state.grup_headers  = gh_val
    # Aktiviteler: JSON'da key string, int'e çevir
    st.session_state.activities    = {int(k): v for k,v in raw_acts.items()}
    # Konsolide eşleştirme kararlarını geri yükle
    if isinstance(data, dict):
        st.session_state.cons_decisions  = data.get("cons_decisions", {})
        st.session_state.cons_rename_map = data.get("cons_rename_map", {})
        st.session_state.cons_confirmed  = data.get("cons_confirmed", False)
    # Eski JSON'lardaki yuvarlama hatalarını temizle:
    # Her özet node (çocuğu olan) için toplamı 100.0000'e normalize et
    _tree = st.session_state.tree
    _id_map = {n["id"]: n for n in _tree}
    for n in _tree:
        kids = [x for x in _tree if x["parent"]==n["id"]]
        if not kids: continue
        total = sum(k["pct"] for k in kids)
        if total == 0: continue
        if abs(total - 100.0) < 0.01:   # küçük hata → normalize et
            factor = 100.0 / total
            for k in kids:
                k["pct"] = round(k["pct"] * factor, 4)
            # Largest remainder düzeltmesi
            s = sum(k["pct"] for k in kids)
            if abs(s - 100.0) > 0.00001:
                diff = round(100.0 - s, 4)
                kids[0]["pct"] = round(kids[0]["pct"] + diff, 4)

# ── Cached ağaç layout + plotly ───────────────────────────────────────────
@st.cache_data(show_spinner=False)
def compute_layout(tj):
    tree = json.loads(tj); id_map = {n["id"]:n for n in tree}
    def depth(nid):
        n = id_map.get(nid)
        return 0 if not n or n["parent"] is None else depth(n["parent"])+1
    def ch(nid): return [n for n in tree if n["parent"]==nid]
    leaves = [n["id"] for n in tree if not ch(n["id"])]
    ly = {lid:i for i,lid in enumerate(leaves)}
    def sy(nid):
        kids = ch(nid)
        return ly[nid] if not kids else sum(sy(k["id"]) for k in kids)/len(kids)
    return {n["id"]:(depth(n["id"]),sy(n["id"])) for n in tree}

@st.cache_data(show_spinner=False)
def build_figure(tj, sel, bedel):
    tree = json.loads(tj); id_map = {n["id"]:n for n in tree}
    pos = compute_layout(tj)
    def ch(nid): return [n for n in tree if n["parent"]==nid]
    def gp(nid):
        n = id_map.get(nid)
        if not n: return 0.0
        if n["parent"] is None: return n["pct"]/100
        sibs = [x for x in tree if x["parent"]==n["parent"]]
        st_ = sum(s["pct"] for s in sibs)
        return gp(n["parent"]) * (n["pct"]/st_ if st_ else 0)
    def ns(nid):
        kids = ch(nid)
        if not kids: return "ok"
        t = sum(k["pct"] for k in kids)
        return "ok" if abs(t-100)<0.05 else ("over" if t>100 else "under")
    ex,ey = [],[]
    for n in tree:
        if n["parent"] is not None and n["parent"] in pos:
            x0,y0=pos[n["parent"]]; x1,y1=pos[n["id"]]
            ex+=[x0,x0,x1,None]; ey+=[y0,y1,y1,None]
    nx,ny,fc,bc,texts,hovers,cdata = [],[],[],[],[],[],[]
    for n in tree:
        x,y = pos[n["id"]]; nx.append(x); ny.append(y); cdata.append(n["id"])
        st_ = ns(n["id"])
        fc.append("#1f6feb" if n["id"]==sel else "#3d1515" if st_!="ok" else "#21262d")
        bc.append("#58a6ff" if n["id"]==sel else
                  "#f85149" if st_=="over" else "#d29922" if st_=="under" else "#484f58")
        short = n["name"] if len(n["name"])<=14 else n["name"][:13]+"…"
        texts.append(f"<b>{short}</b><br>%{n['pct']:.2f}")
        hovers.append(f"<b>{n['name']}</b><br>%{n['pct']:.4f}<br>{gp(n['id'])*bedel:,.2f} TL")
    lc = max(len([n for n in tree if not ch(n["id"])]),1)
    # Node yazıları: her zaman sağ-orta — üst üste binme yok, başlıkla çakışmaz
    return go.Figure(
        data=[go.Scatter(x=ex,y=ey,mode="lines",line=dict(color="#30363d",width=2),hoverinfo="none"),
              go.Scatter(x=nx,y=ny,mode="markers+text",text=texts,
                         textposition="middle right",
                         customdata=cdata,hovertext=hovers,hoverinfo="text",
                         marker=dict(size=22,color=fc,line=dict(color=bc,width=2.5)),
                         textfont=dict(size=11,color="#c9d1d9",family="Inter"))],
        layout=go.Layout(paper_bgcolor="#0d1117",plot_bgcolor="#0d1117",showlegend=False,
            height=max(260,lc*62+100),
            margin=dict(l=30,r=30,t=20,b=20),
            xaxis=dict(showgrid=False,zeroline=False,showticklabels=False,
                       fixedrange=True,
                       range=[min(nx)-0.5, max(nx)+2.5] if nx else None),
            yaxis=dict(showgrid=False,zeroline=False,showticklabels=False,
                       fixedrange=True,
                       range=[min(ny)-0.8, max(ny)+0.8] if ny else None),
            hovermode="closest"))

@st.cache_data(show_spinner=False)
def build_table_df(tj, bedel):
    tree = json.loads(tj); id_map = {n["id"]:n for n in tree}
    def ch(nid): return [n for n in tree if n["parent"]==nid]
    def gp(nid):
        n = id_map.get(nid)
        if not n: return 0.0
        if n["parent"] is None: return n["pct"]/100
        sibs = [x for x in tree if x["parent"]==n["parent"]]
        st_ = sum(s["pct"] for s in sibs)
        return gp(n["parent"]) * (n["pct"]/st_ if st_ else 0)
    def ns(nid):
        kids = ch(nid)
        if not kids: return "Tamam"
        t = sum(k["pct"] for k in kids)
        return "Tamam" if abs(t-100)<0.05 else ("Hata: Fazla" if t>100 else "Uyarı: Eksik")
    rows,ids = [],[]
    def flat(nid, d=0):
        n = id_map.get(nid)
        if not n: return
        pad="　"*d; icon="🔷" if d==0 else("📁" if ch(nid) else "📄")
        rows.append({"  Öğe Tanımı":f"{pad}{icon} {n['name']}",
                     "Ağırlık (%)":n["pct"],
                     "Bütçe (TL)":round(gp(nid)*bedel,2),
                     "Durum":ns(nid)})
        ids.append(nid)
        for kid in ch(nid): flat(kid["id"],d+1)
    for r in [n for n in tree if n["parent"] is None]: flat(r["id"])
    return pd.DataFrame(rows), ids

# ── Excel parse ───────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_excel_new(file_bytes, sheet_name):
    xl = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    raw = []
    for _, row in xl.iterrows():
        no   = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
        if not no or no=="nan" or not name or name=="nan": continue
        try: float(no.split(".")[0])
        except: continue
        d3 = None if (len(row)<=3 or pd.isna(row.iloc[3])) else float(row.iloc[3])
        d4 = None if (len(row)<=4 or pd.isna(row.iloc[4])) else float(row.iloc[4])
        g1 = str(row.iloc[5]).strip() if len(row)>5 and pd.notna(row.iloc[5]) else ""
        g2 = str(row.iloc[6]).strip() if len(row)>6 and pd.notna(row.iloc[6]) else ""
        g3 = str(row.iloc[7]).strip() if len(row)>7 and pd.notna(row.iloc[7]) else ""
        raw.append({"no":no,"name":name,"d3":d3,"d4":d4,"depth":no.count("."),
                    "grup1":g1,"grup2":g2,"grup3":g3})
    all_nos = {r["no"] for r in raw}
    for r in raw:
        r["has_children"] = any(o.startswith(r["no"]+".") for o in all_nos)
    for r in raw:
        if r["has_children"]:
            r["kind"] = "ana_grup" if r["d3"] is None else "ara_grup"
        else:
            r["kind"] = "aktivite"
    return raw

def make_template_excel(grup_headers):
    buf = io.BytesIO()
    h1,h2,h3 = (grup_headers+["Grup 1","Grup 2","Grup 3"])[:3]
    sample = [
        ("","1","RUHSAT, PROJE, SATIŞ İŞLERİ",None,0.0500,"","",""),
        ("","1.1","Ruhsat Projelerinin Hazırlanması",1.000,0.0150,"ABC İnşaat","Proje",""),
        ("","1.2","Ruhsatların Alınması",1.000,0.0100,"ABC İnşaat","Proje",""),
        ("","1.5","Uygulama Projelerinin Hazırlanması",1.000,0.0150,"XYZ Mühendislik","Proje",""),
        ("","1.5.1","Genel Vaziyet Planı",0.050,None,"XYZ Mühendislik","Proje","Öncelikli"),
        ("","1.5.2","Mimari Projeler",0.200,None,"XYZ Mühendislik","Proje",""),
        ("","2","İMALAT İŞLERİ",None,0.9000,"","",""),
        ("","2.1","KAZI VE ZEMİN İŞLERİ",1.000,0.0180,"ABC İnşaat","İnşaat",""),
        ("","2.1.1","Kazı ve Takviye İşleri",0.700,None,"ABC İnşaat","İnşaat",""),
        ("","2.1.2","Zemin İyileştirme",0.300,None,"ZEM Zemin","İnşaat",""),
    ]
    df = pd.DataFrame(sample, columns=[
        "A (Boş)","B (No)","C (Kalem Adı)",
        "D (Alt Pursantaj 0-1)","E (Genel Pursantaj 0-1)",
        f"F ({h1})", f"G ({h2})", f"H ({h3})"
    ])
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="BLOK")
    return buf.getvalue()

# ── Aktivite veri birleştirici (filtre için) ──────────────────────────────
def get_wbs_levels(nid):
    """Kökten node'a giden yolu liste olarak döner: [root, l1, l2, ..., node]"""
    parts = []
    cur = nid
    while cur is not None:
        n = get_node(cur)
        if not n: break
        parts.append(n["name"])
        cur = n["parent"]
    return list(reversed(parts))

def build_act_report(acts, bedel, grup_headers):
    """
    Her satır için:
      - WBS seviye sütunları (L0, L1, L2...): WBS ağacındaki ata adları
      - Aktivite hiyerarşi sütunları (Ana Grup, Ara Grup 1, Ara Grup 2...):
          no derinliğine göre dinamik — max derinlik-1 kadar Ara Grup sütunu

    Kural:
      no="1"     → Ana Grup=boş,    Ara Grup *=boş,  Aktivite Adı=bu satırın adı
      no="1.5"   → Ana Grup=ata(1), Ara Grup *=boş,  Aktivite Adı=bu satırın adı
      no="1.5.2" → Ana Grup=ata(1), Ara Grup 1=ata(1.5), Ara Grup 2=boş, Aktivite Adı=bu satırın adı
      no="2.1.3.1"→ Ana Grup=ata(2), Ara Grup 1=ata(2.1), Ara Grup 2=ata(2.1.3), Aktivite Adı=bu satırın adı

    Returns: (rows_list, max_wbs_depth, max_act_hier_depth)
    """
    h1,h2,h3 = (grup_headers+["Grup 1","Grup 2","Grup 3"])[:3]

    # ── Pass 1: maksimum derinlikleri hesapla ─────────────────────────
    max_wbs_depth = 1
    for nid in acts.keys():
        d = len(get_wbs_levels(nid))
        if d > max_wbs_depth: max_wbs_depth = d

    max_act_depth = 1   # aktivite no'sunun en fazla kaç nokta içerdiği
    for act_list in acts.values():
        if not isinstance(act_list, list): continue
        for r in act_list:
            d = len(r.get("no","").split("."))
            if d > max_act_depth: max_act_depth = d

    # Ara Grup sayısı = max_act_depth - 2  (Ana Grup + Ara Gruplar + yaprak)
    # En az 1 Ara Grup sütunu her zaman çıkar (boş olsa bile)
    n_ara_grp = max(max_act_depth - 2, 1)
    ara_grp_cols = (["Ana Grup"] +
                    [f"Ara Grup {i}" if i > 1 else "Ara Grup" for i in range(1, n_ara_grp+1)])

    # ── Pass 2: satırları oluştur ─────────────────────────────────────
    rows = []
    for nid, act_list in acts.items():
        n = get_node(nid)
        if not n: continue
        nb      = global_pct(nid) * bedel
        bgt_map = compute_act_budgets(act_list, nb)
        wbs_levels = get_wbs_levels(nid)
        no_map  = {r.get("no",""): r for r in act_list}

        for r in act_list:
            d3     = r.get("d3"); d4 = r.get("d4")
            no     = r.get("no","")
            parts  = no.split(".")
            depth  = len(parts)   # 1=ana grup, 2=ara veya aktivite, 3+=derin

            act_bgt = bgt_map.get(no, 0.0)
            kind_tr = {"ana_grup":"Ana Grup","ara_grup":"Ara Grup",
                       "aktivite":"Aktivite"}.get(r["kind"],"—")

            # Ana Grup = ata(depth=1).adı  → satır depth=1 ise boş
            # Ara Grup k = ata(depth=k+1).adı → satır depth<=k+1 ise boş
            # Aktivite Adı = her zaman bu satırın kendi adı

            # Aynı ebeveyn altındaki kardeşlerde ara_grup var mı?
            # Varsa bu seviye "ara grup seviyesi" sayılır —
            # çocuksuz aktivite bile olsa Ara Grup sütununda kendi adını gösterir.
            parent_no = ".".join(parts[:-1])  # "" ise kök
            siblings  = [x for x in act_list
                         if x.get("no","") != no
                         and ".".join(x.get("no","").split(".")[:-1]) == parent_no]
            sibling_has_ara = any(x.get("kind") == "ara_grup" for x in siblings)
            # Bu satır bu seviyede ara_grup gibi davranıyor mu?
            is_ara_level = (r.get("kind") == "ara_grup" or
                            (sibling_has_ara and depth >= 2))

            def ata(target_depth):
                """target_depth basamaklı ata no'sunun adını döner.
                   İstisna: satır ara grup seviyesindeyse (is_ara_level)
                   ve target_depth == depth ise kendi adını döner."""
                if target_depth > depth: return ""
                if target_depth == depth:
                    return r.get("name","") if is_ara_level else ""
                anc_no = ".".join(parts[:target_depth])
                anc_r  = no_map.get(anc_no)
                return anc_r.get("name","") if anc_r else ""

            ana_grup = ata(1)
            # Ara Grup boş kalacaksa → Ana Grup adını kopyala
            # (filtrede hiçbir satır düşmesin, bütçe tutarlı kalsın)

            # Ara Grup 1, Ara Grup 2, ... → ata(2), ata(3), ...
            # Ara Grup değerlerini hesapla
            ara_grups_raw = {
                ("Ara Grup" if i==1 else f"Ara Grup {i}"): ata(i+1)
                for i in range(1, n_ara_grp+1)
            }
            # Fallback: her Ara Grup sütunu boşsa → Ana Grup adını yaz
            ara_grups = {}
            for col, val in ara_grups_raw.items():
                if val:
                    ara_grups[col] = val
                else:
                    # Sadece bu satır depth>=2 ve ana_grup doluysa fallback uygula
                    ara_grups[col] = ana_grup if (depth >= 2 and ana_grup) else ""

            row = {
                "wbs_id":         nid,
                "WBS Öğesi":      node_path(nid),
                "WBS Adı":        n["name"],
                "WBS Bütçe (TL)": round(nb,2),
                "no":             no,
                "Aktivite No":    "'" + no,
                "Tür":            kind_tr,
                "kind":           r["kind"],
                "Ana Grup":       ana_grup,
                **ara_grups,
                "Aktivite Adı":   r.get("name",""),
                "d3":             d3,
                "d4":             d4,
                "Alt Pay (%)":    round(d3*100,4) if d3 is not None else None,
                "Genel Pay (%)":  round(d4*100,4) if d4 is not None else None,
                "Bütçe (TL)":     round(act_bgt,2),
                h1:               r.get("grup1",""),
                h2:               r.get("grup2",""),
                h3:               r.get("grup3",""),
            }
            # WBS seviye sütunları
            for li in range(max_wbs_depth):
                row[f"L{li}"] = wbs_levels[li] if li < len(wbs_levels) else ""
            rows.append(row)

    return rows, max_wbs_depth, n_ara_grp


# ── Raporlar ──────────────────────────────────────────────────────────────
def build_excel_report(flat, act_rows_data, errors, warnings_list, info_list, bedel, grup_headers):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()

    # --- Açık tema renkleri ---
    H_BG  = "2C5F8A"  # başlık mavi
    H_FG  = "FFFFFF"
    L0_BG = "D6E4F0"  # ana grup
    L1_BG = "EAF4FB"  # ara grup
    L2_BG = "F8FBFF"  # aktivite
    ALT_BG= "F0F7FF"
    RED_BG= "FDECEA"; RED_FG= "C0392B"
    GRN_BG= "E9F7EF"; GRN_FG= "1E8449"
    ORG_BG= "FEF9E7"; ORG_FG= "784212"
    TEXT  = "1A1A2E"
    BORDER_CLR = "B0C4DE"

    thin  = Side(style="thin", color=BORDER_CLR)
    def brd(): return Border(top=thin,left=thin,right=thin,bottom=thin)
    def fill(c): return PatternFill("solid", fgColor=c)
    def font(c=TEXT, bold=False, sz=10): return Font(color=c,bold=bold,size=sz,name="Calibri")
    def aln(h="left",v="center",indent=0): return Alignment(horizontal=h,vertical=v,indent=indent,wrap_text=False)

    wb = __import__("openpyxl").Workbook()

    # ── Sayfa 1: WBS Özet ──────────────────────────────────────────────
    ws = wb.active; ws.title = "WBS Özet"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"WBS Maliyet Özeti  |  Sözleşme Bedeli: {bedel:,.2f} TL"
    ws["A1"].fill  = fill(H_BG); ws["A1"].font = font(H_FG,True,13)
    ws["A1"].alignment = aln("center"); ws.row_dimensions[1].height = 26

    hdrs = ["Öğe Adı","Seviye","Yerel Ağırlık","Genel Pay","Bütçe (TL)","Durum","Aktivite"]
    for ci,h in enumerate(hdrs,1):
        c = ws.cell(2,ci,h)
        c.fill=fill(H_BG); c.font=font(H_FG,True,9)
        c.alignment=aln("center"); c.border=brd()
    ws.row_dimensions[2].height = 18

    bg_map = {0:L0_BG, 1:L1_BG}
    for ri,r in enumerate(flat,3):
        bg  = bg_map.get(r["depth"], L2_BG if ri%2==0 else ALT_BG)
        fw  = r["depth"] <= 1
        ind = r["depth"]
        dur = {"ok":"✅ Tamam","over":"❌ Fazla","under":"⚠️ Eksik"}.get(r["status"],"—")
        act_s = f"{r['act_count']} aktivite" if r["has_acts"] else "—"
        vals  = [r["name"], r["depth"],
                 r["pct_local"]/100,  # decimal → "0.00%" format gösterir doğru
                 r["pct_global"]/100,
                 r["budget"], dur, act_s]
        fmts  = [None,None,"0.00%","0.0000%",'#,##0.00 "TL"',None,None]
        alns  = [aln("left",indent=ind),aln("center"),aln("right"),
                 aln("right"),aln("right"),aln("center"),aln("center")]
        for ci,(v,fmt,al) in enumerate(zip(vals,fmts,alns),1):
            cell = ws.cell(ri,ci,v)
            cell.fill=fill(bg); cell.font=font(TEXT,fw,9 if not fw else 10)
            cell.alignment=al; cell.border=brd()
            if fmt and isinstance(v,(int,float)): cell.number_format=fmt
        ws.row_dimensions[ri].height=15

    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col),default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+3,55)
    ws.column_dimensions["A"].width=50

    # ── Sayfa 2: Aktivite Detay ─────────────────────────────────────────
    if act_rows_data:
        h1,h2,h3 = (grup_headers+["Grup 1","Grup 2","Grup 3"])[:3]
        ws2 = wb.create_sheet("Aktivite Detay")
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A3"

        ws2.merge_cells(f"A1:K1")
        ws2["A1"].value = f"Aktivite Detay Tablosu  |  Sözleşme: {bedel:,.2f} TL"
        ws2["A1"].fill=fill(H_BG); ws2["A1"].font=font(H_FG,True,13)
        ws2["A1"].alignment=aln("center"); ws2.row_dimensions[1].height=26

        ah = ["WBS Öğesi","WBS Bütçe (TL)","Aktivite No","Aktivite Adı",
              "Tür","Alt Pay","Genel Pay","Bütçe (TL)",h1,h2,h3]
        for ci,h in enumerate(ah,1):
            c=ws2.cell(2,ci,h)
            c.fill=fill(H_BG); c.font=font(H_FG,True,9)
            c.alignment=aln("center"); c.border=brd()
        ws2.row_dimensions[2].height=18

        kind_bg = {"Ana Grup":L0_BG,"Ara Grup":L1_BG,"Aktivite":L2_BG}
        for ri,row in enumerate(act_rows_data,3):
            bg  = kind_bg.get(row["Tür"],ALT_BG) if ri%2==0 else kind_bg.get(row["Tür"],L2_BG)
            d3v = row["d3"] if row["d3"] is not None else None
            d4v = row["d4"] if row["d4"] is not None else None
            vals=[row["WBS Öğesi"], row["WBS Bütçe (TL)"],
                  row["Aktivite No"], row["Aktivite Adı"], row["Tür"],
                  d3v, d4v, row["Bütçe (TL)"],
                  row.get(h1,""), row.get(h2,""), row.get(h3,"")]
            fmts=[None,'#,##0.00 "TL"',None,None,None,"0.00%","0.0000%",'#,##0.00 "TL"',None,None,None]
            for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
                cell=ws2.cell(ri,ci,v)
                cell.fill=fill(bg); cell.font=font(TEXT,row["kind"]!="aktivite",9)
                cell.alignment=aln("right") if isinstance(v,(int,float)) and v is not None else aln("left")
                cell.border=brd()
                if fmt and isinstance(v,(int,float)) and v is not None:
                    cell.number_format=fmt
            ws2.row_dimensions[ri].height=15

        for col in ws2.columns:
            mx=max((len(str(c.value or "")) for c in col),default=8)
            ws2.column_dimensions[get_column_letter(col[0].column)].width=min(mx+3,50)
        ws2.column_dimensions["D"].width=50

    # ── Sayfa 3: Kontrol ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Kontrol Raporu")
    ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:B1")
    ws3["A1"].value="Doğruluk Kontrol Raporu"
    ws3["A1"].fill=fill(H_BG); ws3["A1"].font=font(H_FG,True,12)
    ws3["A1"].alignment=aln("center"); ws3.row_dimensions[1].height=24
    for ci,h in enumerate(["Tür","Mesaj"],1):
        ws3.cell(2,ci,h).fill=fill(H_BG); ws3.cell(2,ci,h).font=font(H_FG,True)
        ws3.cell(2,ci,h).border=brd()
    sty_map={"HATA":(RED_BG,RED_FG),"UYARI":(ORG_BG,ORG_FG),"BİLGİ":(GRN_BG,GRN_FG)}
    for ri,row in enumerate([{"Tür":"HATA","Mesaj":e} for e in errors]+
                             [{"Tür":"UYARI","Mesaj":w} for w in warnings_list]+
                             [{"Tür":"BİLGİ","Mesaj":i} for i in info_list],3):
        bg2,fg2=sty_map.get(row["Tür"],(L2_BG,TEXT))
        ws3.cell(ri,1,row["Tür"]).fill=fill(bg2); ws3.cell(ri,1).font=font(fg2,True)
        ws3.cell(ri,1).alignment=aln("center"); ws3.cell(ri,1).border=brd()
        ws3.cell(ri,2,row["Mesaj"]).fill=fill(bg2); ws3.cell(ri,2).font=font(fg2)
        ws3.cell(ri,2).border=brd()
    ws3.column_dimensions["A"].width=12; ws3.column_dimensions["B"].width=80

    wb.save(buf); return buf.getvalue()

def build_html_report(flat, act_rows_data, errors, warnings_list, info_list, bedel):
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    kind_clr={"Ana Grup":("#D6E4F0","#1a5276"),"Ara Grup":("#EAF4FB","#1a5276"),"Aktivite":("#FDFEFE","#2c3e50")}

    wbs_html=""
    for r in flat:
        bg={"0":"#D6E4F0","1":"#EAF4FB"}.get(str(r["depth"]),"#F8FBFF" if r["depth"]%2==0 else "#FDFEFE")
        fw="700" if r["depth"]<=1 else "400"
        pad=r["depth"]*18
        dur={"ok":"✅ Tamam","over":"❌ Fazla","under":"⚠️ Eksik"}.get(r["status"],"—")
        wbs_html+=f'<tr style="background:{bg}"><td style="padding-left:{pad+8}px;font-weight:{fw}">{r["name"]}</td><td style="text-align:center">{r["depth"]}</td><td style="text-align:right">{r["pct_local"]:.2f}%</td><td style="text-align:right">{r["pct_global"]:.4f}%</td><td style="text-align:right;font-weight:600;color:#1a5276">{r["budget"]:,.2f} TL</td><td style="text-align:center">{dur}</td><td style="text-align:center;color:#666">{r["act_count"] if r["has_acts"] else "—"}</td></tr>\n'

    act_html=""
    for row in act_rows_data:
        bg,fg=kind_clr.get(row["Tür"],("#FDFEFE","#2c3e50"))
        dot={"Ana Grup":"🔵","Ara Grup":"🟢","Aktivite":"🟠"}.get(row["Tür"],"●")
        ap=f"{row['Alt Pay (%)']:.2f}%" if row["Alt Pay (%)"] is not None else "—"
        gp2=f"{row['Genel Pay (%)']:.4f}%" if row["Genel Pay (%)"] is not None else "—"
        act_html+=f'<tr style="background:{bg};color:{fg}"><td>{row["WBS Öğesi"]}</td><td style="text-align:right">{row["WBS Bütçe (TL)"]:,.2f} TL</td><td>{row["Aktivite No"]}</td><td>{row["Aktivite Adı"]}</td><td>{dot} {row["Tür"]}</td><td style="text-align:right">{ap}</td><td style="text-align:right">{gp2}</td><td style="text-align:right;font-weight:700;color:#1a5276">{row["Bütçe (TL)"]:,.2f} TL</td></tr>\n'

    ctrl=""
    for e in errors:    ctrl+=f'<div style="background:#FDECEA;border-left:4px solid #C0392B;padding:.4rem .8rem;margin:.2rem 0;border-radius:4px;color:#922B21">❌ {e}</div>'
    for w in warnings_list: ctrl+=f'<div style="background:#FEF9E7;border-left:4px solid #F39C12;padding:.4rem .8rem;margin:.2rem 0;border-radius:4px;color:#784212">⚠️ {w}</div>'
    for i in info_list: ctrl+=f'<div style="background:#E9F7EF;border-left:4px solid #27AE60;padding:.4rem .8rem;margin:.2rem 0;border-radius:4px;color:#1E8449">✅ {i}</div>'

    act_sec=""
    if act_rows_data:
        act_sec=f"""<h2>📋 Aktivite Detay Tablosu</h2>
<table><thead><tr><th>WBS Öğesi</th><th>WBS Bütçe</th><th>No</th><th>Aktivite Adı</th><th>Tür</th><th>Alt %</th><th>Genel %</th><th>Bütçe (TL)</th></tr></thead><tbody>{act_html}</tbody></table>"""

    return f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8"><title>Pursantaj Raporu</title>
<style>
@media print{{body{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}.no-print{{display:none}}}}
*{{margin:0;padding:0;box-sizing:border-box;font-family:'Segoe UI',Arial,sans-serif}}
body{{background:#fff;color:#2c3e50;padding:1.5rem 2rem;font-size:11px}}
h1{{text-align:center;color:#1a5276;font-size:1.5rem;margin-bottom:.3rem}}
h2{{color:#1a5276;font-size:1rem;margin:1.2rem 0 .5rem;border-bottom:2px solid #2C5F8A;padding-bottom:.3rem}}
.meta{{text-align:center;color:#666;font-size:.8rem;margin-bottom:1rem}}
.score-row{{display:flex;gap:.8rem;margin:.8rem 0}}
.score-box{{flex:1;background:#EBF5FB;border:1px solid #AED6F1;border-radius:6px;padding:.5rem .8rem;text-align:center}}
.score-box .v{{font-size:1.1rem;font-weight:700;color:#1a5276}}
.score-box .l{{font-size:.7rem;color:#666}}
table{{width:100%;border-collapse:collapse;font-size:10.5px;margin-bottom:.5rem}}
th{{background:#2C5F8A;color:#fff;padding:.4rem .6rem;text-align:left;font-weight:600;border:1px solid #AED6F1;white-space:nowrap}}
td{{padding:.35rem .6rem;border:1px solid #D6EAF8}}
.print-btn{{position:fixed;bottom:1.5rem;right:1.5rem;background:#2C5F8A;color:#fff;border:none;padding:.6rem 1.3rem;border-radius:7px;font-size:.9rem;cursor:pointer;font-weight:600;box-shadow:0 3px 10px rgba(0,0,0,.2)}}
.print-btn:hover{{background:#1a4a6e}}
</style></head><body>
<button class="print-btn no-print" onclick="window.print()">🖨️ PDF Olarak Kaydet</button>
<h1>Pursantaj Yönetim Sistemi — Rapor</h1>
<div class="meta">Sözleşme Bedeli: <b style="color:#1a5276">{bedel:,.2f} TL</b> &nbsp;|&nbsp; {now}</div>
<div class="score-row">
  <div class="score-box"><div class="v">{len(flat)}</div><div class="l">Toplam Öğe</div></div>
  <div class="score-box"><div class="v">{sum(1 for r in flat if r["is_leaf"])}</div><div class="l">Yaprak Öğe</div></div>
  <div class="score-box"><div class="v" style="color:{'#C0392B' if errors else '#1E8449'}">{len(errors)}</div><div class="l">Hata</div></div>
  <div class="score-box"><div class="v" style="color:{'#D35400' if warnings_list else '#1E8449'}">{len(warnings_list)}</div><div class="l">Uyarı</div></div>
  <div class="score-box"><div class="v">{len(act_rows_data)}</div><div class="l">Aktivite Satırı</div></div>
</div>
{ctrl}
<h2>🌳 WBS Maliyet Özeti</h2>
<table><thead><tr><th>Öğe Adı</th><th>Sev.</th><th>Yerel %</th><th>Genel %</th><th>Bütçe (TL)</th><th>Durum</th><th>Aktivite</th></tr></thead><tbody>{wbs_html}</tbody></table>
{act_sec}
</body></html>""".encode("utf-8")

# ══════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════
st.markdown("""
<h1 style="margin-bottom:0">Pursantaj Yönetim Sistemi</h1>
<div style="height:.5rem"></div>
""", unsafe_allow_html=True)
tab1,tab2,tab3,tab4,tab5 = st.tabs([
    "🌳  WBS Ağacı", "📋  Aktiviteler", "📊  Özet & Rapor", "🔍  Filtre & Analiz", "📈  Grafik"])

# ══════════════════════════════════════════════════════════════════════════
# TAB 1 — WBS Ağacı
# ══════════════════════════════════════════════════════════════════════════
with tab1:
    tj    = json.dumps(st.session_state.tree, sort_keys=True)
    sel   = st.session_state.sel
    bedel = st.session_state.bedel
    sn    = get_node(sel)

    fig = build_figure(tj, sel, bedel)
    clicked = st.plotly_chart(fig, use_container_width=True,
                              on_select="rerun", selection_mode="points", key="tc")
    if clicked and clicked.get("selection",{}).get("points"):
        pt = clicked["selection"]["points"][0]
        cd = pt.get("customdata")
        if cd is not None and int(cd) != sel:
            st.session_state.sel = int(cd); st.rerun()

    df, ids = build_table_df(tj, bedel)
    ev = st.dataframe(df, use_container_width=True, hide_index=True,
                      on_select="rerun", selection_mode="single-row", key="tbl",
                      column_config={
                          "  Öğe Tanımı": st.column_config.TextColumn(width="large"),
                          "Ağırlık (%)":  st.column_config.NumberColumn(format="%.2f %%", width="small"),
                          "Bütçe (TL)":   st.column_config.NumberColumn(format="%,.2f", width="medium"),
                          "Durum":        st.column_config.TextColumn(width="small"),
                      })
    if ev and ev.selection and ev.selection.rows:
        ri = ev.selection.rows[0]
        if ri < len(ids) and ids[ri] != sel:
            st.session_state.sel = ids[ri]; st.rerun()

    fmt = lambda v: f"{v:,.2f} TL".replace(",","X").replace(".",",").replace("X",".")
    st.markdown(f"""
    <div class="mrow">
      <div class="mbox"><div class="lbl">Seçili Öğe</div>
        <div class="val">{sn['name'] if sn else '—'}</div>
        <div class="sub">{'📄 Yaprak' if sn and is_leaf(sel) else '📁 Özet öğe'}</div></div>
      <div class="mbox"><div class="lbl">Seçili Bütçe</div>
        <div class="val">{fmt(global_pct(sel)*bedel) if sn else '—'}</div>
        <div class="sub">Pay: %{global_pct(sel)*100:.4f}</div></div>
      <div class="mbox"><div class="lbl">Toplam Sözleşme</div>
        <div class="val">{fmt(bedel)}</div>
        <div class="sub">{sum(1 for n in st.session_state.tree if is_leaf(n["id"]))} yaprak öğe</div></div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")

    with st.form("ctrl", border=False):
        cb,ca,cn = st.columns([1.1,1,1.2])
        with cb:
            bs = st.text_input("Sözleşme Bedeli (TL)",
                 value=f"{bedel:,.2f}".replace(",","X").replace(".",",").replace("X","."))
        with ca:
            max_av = round(100.0-sibling_total_excl(sel),6) if sn and sn["parent"] is not None else 100.0
            ps = st.text_input(f"Ağırlık (%) [max %{max_av:.4f}]", value=str(sn["pct"]) if sn else "0")
        with cn:
            nv = st.text_input("Öğe Adı", value=sn["name"] if sn else "")
        ap = st.form_submit_button("✔ Uygula", type="primary")
    if ap:
        changed = False
        try:
            p = float(bs.replace(".","").replace(",","."))
            if abs(p-st.session_state.bedel)>0.001: st.session_state.bedel=p; changed=True
        except: pass
        if sn:
            try:
                np_ = max(0.0,min(round(float(ps.replace(",",".")),6),max_av))
                if abs(np_-sn["pct"])>1e-9: sn["pct"]=np_; changed=True
            except: pass
            if nv and nv!=sn["name"]: sn["name"]=nv; changed=True
        if changed: st.rerun()

    c1,c2,c3,c4,c5 = st.columns(5)
    with c1:
        if st.button("➕ Alt Öğe"):
            nid=add_child(sel); st.session_state.sel=nid; st.rerun()
    with c2:
        if st.button("↔️ Kardeş Öğe"):
            nid=add_sibling(sel); st.session_state.sel=nid; st.rerun()
    with c3:
        if st.button("📋 Kopyala / Çoğalt"):
            st.session_state.copy_mode=True; st.rerun()
    with c4:
        if st.button("🗑️ Öğeyi Sil"):
            n=get_node(sel)
            if n and n["parent"] is not None:
                pid=n["parent"]; delete_subtree(sel)
                st.session_state.sel=pid; auto_distribute(pid); st.rerun()
            else: st.warning("Kök öğe silinemez.")
    with c5:
        st.download_button("📥 JSON İndir", data=export_json(),
                           file_name="wbs_proje.json", mime="application/json")

    # JSON Yükle
    with st.expander("📂 JSON Dosyası Yükle (Ağaç + Aktiviteler)", expanded=False):
        st.caption("Daha önce dışa aktardığınız `wbs_proje.json` dosyasını yükleyin.")
        json_file = st.file_uploader("wbs_proje.json", type=["json"], key="json_upload")
        if json_file:
            try:
                data = json.loads(json_file.read().decode("utf-8"))
                # Hem eski (liste) hem yeni (dict) formatı destekle
                if isinstance(data, list):
                    tree_data = data
                    act_count = 0
                elif isinstance(data, dict) and "tree" in data:
                    tree_data = data["tree"]
                    act_count = sum(len(v) for v in data.get("activities",{}).values())
                else:
                    st.error("❌ Geçersiz JSON formatı."); tree_data = None
                if tree_data is not None:
                    if not isinstance(tree_data, list) or not tree_data:
                        st.error("❌ Geçersiz format.")
                    else:
                        roots = [n for n in tree_data if n["parent"] is None]
                        st.info(f"**Önizleme:** {len(tree_data)} WBS öğesi, {len(roots)} kök, {act_count} aktivite satırı")
                    if tree_data is not None:
                        c_ok,c_cancel=st.columns(2)
                        with c_ok:
                            if st.button("✔ Yükle ve Uygula", type="primary"):
                                import_json(data)
                                st.success("✅ Proje ve aktiviteler yüklendi!"); st.rerun()
                        with c_cancel:
                            st.info("İptal için dosyayı kaldırın.")
            except Exception as e:
                st.error(f"❌ {e}")

    # Kopyala diyaloğu
    if st.session_state.get("copy_mode") and sn:
        st.markdown(f"""<div class="copy-box">
          <b style="color:#d29922">📋 Kopyala / Çoğalt — {sn['name']}</b><br>
          <span style="font-size:.8rem;color:#8b949e">Tüm çocukları ve aktiviteleriyle kopyalanacak.</span>
        </div>""", unsafe_allow_html=True)
        with st.form("copy_form", border=False):
            cc1,cc2,cc3 = st.columns([1,1.5,1])
            with cc1: n_copies = st.number_input("Kopya sayısı",min_value=1,max_value=20,value=1,step=1)
            with cc2: base_name = st.text_input("Temel isim (boş=otomatik)",value="")
            with cc3: copy_acts = st.checkbox("Aktiviteleri de kopyala",value=True)
            ok_c,can_c = st.columns(2)
            with ok_c: ok = st.form_submit_button("✔ Oluştur",type="primary")
            with can_c: cancel = st.form_submit_button("✕ İptal")
        if ok:
            pid = sn["parent"]
            existing = [n["name"] for n in st.session_state.tree if n["parent"]==pid and n["id"]!=sel]
            for i in range(1,int(n_copies)+1):
                if base_name.strip():
                    cname = f"{base_name.strip()}"+(f"_{i}" if n_copies>1 else "")
                else:
                    suf=2; cname=f"{sn['name']}_{suf}"
                    while cname in existing: suf+=1; cname=f"{sn['name']}_{suf}"
                    existing.append(cname)
                new_id = copy_subtree(sel, pid, cname)
                if not copy_acts and new_id is not None:
                    def clr(nid):
                        st.session_state.activities.pop(nid,None)
                        for k in children_of(nid): clr(k["id"])
                    clr(new_id)
            if pid is not None: auto_distribute(pid)
            st.session_state.copy_mode=False; st.rerun()
        if cancel: st.session_state.copy_mode=False; st.rerun()

# ══════════════════════════════════════════════════════════════════════════
# TAB 2 — Aktiviteler
# ══════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("#### 📋 Aktivite Yönetimi")
    gh = st.session_state.grup_headers

    # Grup başlığı ayarları
    with st.expander("⚙️ Grup Sütun Başlıkları", expanded=st.session_state.show_grup_headers):
        gc1,gc2,gc3,gc4 = st.columns([1,1,1,1])
        with gc1: g1n = st.text_input("Grup 1 Başlığı", value=gh[0], key="gh1")
        with gc2: g2n = st.text_input("Grup 2 Başlığı", value=gh[1], key="gh2")
        with gc3: g3n = st.text_input("Grup 3 Başlığı", value=gh[2], key="gh3")
        with gc4:
            st.write("")
            if st.button("💾 Başlıkları Kaydet"):
                st.session_state.grup_headers = [g1n,g2n,g3n]
                gh = [g1n,g2n,g3n]; st.rerun()

    h1,h2,h3 = gh

    # Şablon
    with st.expander("📄 Excel Şablonu & Format Rehberi", expanded=False):
        st.markdown(f"""
| Sütun | İçerik | Açıklama |
|-------|--------|----------|
| A | *(boş)* | |
| B | **No** | `1` / `1.1` / `1.5.2` |
| C | **Kalem Adı** | |
| D | **Alt Pursantaj** | Kardeşler arası 0–1. Ana gruplarda boş. |
| E | **Genel Pursantaj** | Proje geneli 0–1. Aktivitelerde boş. |
| F | **{h1}** | Grup 1 etiketi |
| G | **{h2}** | Grup 2 etiketi |
| H | **{h3}** | Grup 3 etiketi |
        """)
        st.download_button("📥 Şablon İndir", data=make_template_excel(gh),
                           file_name="aktivite_sablonu.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")

    # Öğe seçici
    all_opts = []
    def collect_all(nid, d=0):
        n = get_node(nid)
        if not n: return
        icon = "🔷" if d==0 else("📁" if not is_leaf(nid) else "📄")
        all_opts.append((nid, f"{icon}{'  '*d} {n['name']}"))
        for ch in children_of(nid): collect_all(ch["id"],d+1)
    for root in [n for n in st.session_state.tree if n["parent"] is None]: collect_all(root["id"])
    node_ids=[x[0] for x in all_opts]; node_labels=[x[1] for x in all_opts]
    def_idx=node_ids.index(st.session_state.sel) if st.session_state.sel in node_ids else 0
    sel_label=st.selectbox("Aktivite eklenecek öğe",node_labels,index=def_idx,key="act_sel")
    sel_act_id=node_ids[node_labels.index(sel_label)]
    san=get_node(sel_act_id)
    bgt=global_pct(sel_act_id)*st.session_state.bedel
    st.markdown(f"""
    <div class="mrow">
      <div class="mbox"><div class="lbl">Seçili Öğe</div><div class="val">{san['name'] if san else '—'}</div>
        <div class="sub">Pay: %{global_pct(sel_act_id)*100:.4f}</div></div>
      <div class="mbox"><div class="lbl">Öğe Bütçesi</div><div class="val">{bgt:,.2f} TL</div>
        <div class="sub">Sözleşme: {st.session_state.bedel:,.2f} TL</div></div>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")
    uc1,uc2=st.columns([2,1])
    with uc1: uploaded=st.file_uploader("Excel Dosyası (.xlsx)",type=["xlsx","xls"],key="af")
    with uc2: sheet_nm=st.text_input("Sayfa Adı",value="BLOK",key="sh")

    if uploaded:
        fb=uploaded.read()
        try: act_rows=parse_excel_new(fb,sheet_nm)
        except Exception as e: st.error(f"Excel okunamadı: {e}"); act_rows=[]

        if act_rows:
            ana_grp=[r for r in act_rows if r["kind"]=="ana_grup"]
            ara_grp=[r for r in act_rows if r["kind"]=="ara_grup"]
            akt    =[r for r in act_rows if r["kind"]=="aktivite"]
            _sel_nb=global_pct(sel_act_id)*st.session_state.bedel
            bgt_map=compute_act_budgets(act_rows, _sel_nb)  # seçili öğe bütçesi

            st.markdown(f"**{len(act_rows)} kalem** — {len(ana_grp)} ana grup / {len(ara_grp)} ara grup / {len(akt)} aktivite")

            disp=[]
            for r in act_rows:
                pad="　"*r["depth"]
                icon={"ana_grup":"🔵","ara_grup":"🟢","aktivite":"🟠"}.get(r["kind"],"📄")
                act_bgt=bgt_map.get(r["no"],0.0)
                disp.append({
                    "  Aktivite": f"{pad}{icon} {r['no']}  {r['name']}",
                    "Tür":        {"ana_grup":"Ana Grup","ara_grup":"Ara Grup","aktivite":"Aktivite"}.get(r["kind"]),
                    "D (Alt %)":  round(r["d3"]*100,4) if r["d3"] is not None else None,
                    "E (Genel %)":round(r["d4"]*100,4) if r["d4"] is not None else None,
                    "Bütçe (TL)": round(act_bgt,2),
                    h1:           r.get("grup1",""),
                    h2:           r.get("grup2",""),
                    h3:           r.get("grup3",""),
                })
            st.dataframe(pd.DataFrame(disp), hide_index=True, use_container_width=True,
                         column_config={
                             "D (Alt %)":  st.column_config.NumberColumn(format="%.4f %%"),
                             "E (Genel %)":st.column_config.NumberColumn(format="%.4f %%"),
                             "Bütçe (TL)": st.column_config.NumberColumn(format="%,.2f"),
                         })

            # Doğrulama soruları
            st.markdown("### ❓ Doğrulama")
            q1=st.radio("**S1** Numara sistemi (B sütunu) hiyerarşiyi doğru temsil ediyor mu?",
                        ["✅ Evet","❌ Hayır"],key="q1",horizontal=True)
            q2=st.radio("**S2** D sütunu kardeşler arası göreli ağırlık (0–1)?",
                        ["✅ Evet","❌ Hayır"],key="q2",horizontal=True)
            q3=st.radio("**S3** E sütunu proje geneli mutlak oran (0–1)?",
                        ["✅ Evet","❌ Hayır"],key="q3",horizontal=True)

            all_ok = all("Evet" in q for q in [q1,q2,q3])
            if all_ok:
                if st.button("💾 Onaylayarak Kaydet", type="primary"):
                    st.session_state.activities[sel_act_id]=act_rows
                    st.success(f"✅ {len(act_rows)} kalem kaydedildi!")
            else:
                st.error("⛔ Sorunları çözün, ardından kaydedin.")

    # Kaydedilmiş aktiviteler
    st.markdown("---")
    st.markdown("#### 💾 Kaydedilmiş Aktiviteler")
    saved=st.session_state.activities
    if not saved: st.info("Henüz aktivite eklenmedi.")
    else:
        for nid,acts in saved.items():
            n=get_node(nid); nm=n["name"] if n else f"Öğe {nid}"
            nb=global_pct(nid)*st.session_state.bedel
            _nb=nb
            bgt_m=compute_act_budgets(acts,_nb)
            gc=sum(1 for a in acts if a["kind"]!="aktivite")
            ac=sum(1 for a in acts if a["kind"]=="aktivite")
            full_path=node_path(nid) if n else nm
            exp_title=(f"{'📁' if not is_leaf(nid) else '📄'} "
                       f"{full_path}  —  {len(acts)} kalem ({gc} grup / {ac} akt.)"
                       f"  |  {nb:,.2f} TL")
            with st.expander(exp_title):
                rs=[]
                for r in acts:
                    icon={"ana_grup":"🔵","ara_grup":"🟢","aktivite":"🟠"}.get(r["kind"],"📄")
                    rs.append({"  Kalem":f"{'　'*r['depth']}{icon} {r['no']}  {r['name']}",
                               "Tür":{"ana_grup":"Ana Grup","ara_grup":"Ara Grup","aktivite":"Aktivite"}.get(r["kind"]),
                               "Bütçe (TL)":round(bgt_m.get(r["no"],0),2),
                               h1:r.get("grup1",""), h2:r.get("grup2",""), h3:r.get("grup3","")})
                st.dataframe(pd.DataFrame(rs),hide_index=True,use_container_width=True,
                             column_config={"Bütçe (TL)":st.column_config.NumberColumn(format="%,.2f")})
                if st.button(f"🗑️ Sil",key=f"del_{nid}"):
                    del st.session_state.activities[nid]; st.rerun()

# ══════════════════════════════════════════════════════════════════════════
# TAB 3 — Özet & Rapor
# ══════════════════════════════════════════════════════════════════════════
with tab3:
    bedel = st.session_state.bedel
    acts  = st.session_state.activities
    gh    = st.session_state.grup_headers

    def full_flat():
        result=[]
        def walk(nid, depth=0):
            n=get_node(nid)
            if not n: return
            kids=children_of(nid); gp_=global_pct(nid)
            al=acts.get(nid,[])
            result.append({
                "nid":nid,"depth":depth,"name":n["name"],
                "pct_local":n["pct"],"pct_global":gp_*100,
                "budget":gp_*bedel,"status":node_status(nid),
                "is_leaf":len(kids)==0,"has_acts":nid in acts,
                "act_count":len(al),"act_grps":sum(1 for a in al if a["kind"]!="aktivite"),
                "act_acts":sum(1 for a in al if a["kind"]=="aktivite"),
            })
            for k in kids: walk(k["id"],depth+1)
        for r in [n for n in st.session_state.tree if n["parent"] is None]: walk(r["id"])
        return result

    flat = full_flat()
    # Cache'den al (Tab4 ile aynı hesabı paylaş)
    _cache_key_oz = (id(acts), bedel, tuple(gh))
    if st.session_state.get("_act_cache_key") == _cache_key_oz:
        act_rows_data = st.session_state["_act_cache_rows"]
    else:
        act_rows_data, _, _n_ara = build_act_report(acts, bedel, gh)

    # Doğruluk kontrolleri
    errors,warnings_list,info_list=[],[],[]
    for n in st.session_state.tree:
        kids=children_of(n["id"])
        if kids:
            t=sum(k["pct"] for k in kids)
            if abs(t-100)>0.05:
                (errors if abs(t-100)>1 else warnings_list).append(
                    f"**{n['name']}** kardeş toplamı %{t:.2f}")
    roots=[n for n in st.session_state.tree if n["parent"] is None]
    if abs(sum(global_pct(r["id"])*100 for r in roots)-100)>0.05:
        errors.append("Kök toplam %100 değil")
    else: info_list.append("Kök toplam %100 ✅")
    no_act=[r for r in flat if r["is_leaf"] and not r["has_acts"]]
    if no_act: warnings_list.append(f"{len(no_act)} yaprak öğede aktivite yok")
    else: info_list.append("Tüm yaprak öğelerde aktivite tanımlı ✅")

    # Skorlar
    st.markdown("### 🔍 Doğruluk Kontrolü")
    sm1,sm2,sm3,sm4=st.columns(4)
    sm1.metric("Genel Durum","✅ Hazır" if not errors and not warnings_list else ("⚠️ Uyarı" if not errors else "❌ Hata"))
    sm2.metric("Hata",f"{'🔴' if errors else '🟢'} {len(errors)}")
    sm3.metric("Uyarı",f"{'🟡' if warnings_list else '🟢'} {len(warnings_list)}")
    sm4.metric("Aktivite",f"{len(act_rows_data)} satır")
    for e in errors: st.error(f"❌ {e}")
    for w in warnings_list: st.warning(f"⚠️ {w}")
    for i in info_list: st.success(f"✅ {i}")

    st.markdown("---")
    # Metrikler
    st.markdown("### 📈 Proje Özeti")
    leaf_count=sum(1 for r in flat if r["is_leaf"])
    act_cov=sum(1 for r in flat if r["is_leaf"] and r["has_acts"])
    st.markdown(f"""
    <div class="mrow">
      <div class="mbox"><div class="lbl">Sözleşme Bedeli</div><div class="val">{bedel:,.2f} TL</div></div>
      <div class="mbox"><div class="lbl">WBS Öğe</div><div class="val">{len(flat)}</div>
        <div class="sub">{leaf_count} yaprak / {len(flat)-leaf_count} özet</div></div>
      <div class="mbox"><div class="lbl">Aktivite Kapsama</div><div class="val">{act_cov}/{leaf_count}</div>
        <div class="sub">{len(act_rows_data)} toplam satır</div></div>
      <div class="mbox"><div class="lbl">Derinlik</div><div class="val">{max((r["depth"] for r in flat),default=0)+1} seviye</div></div>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 🌳 Tam Hiyerarşik Maliyet Tablosu")
    rdf=pd.DataFrame([{
        "  Öğe Tanımı":"　"*r["depth"]+("🔷" if r["depth"]==0 else("📁" if not r["is_leaf"] else "📄"))+" "+r["name"],
        "Seviye":r["depth"],"Yerel %":round(r["pct_local"],4),"Genel %":round(r["pct_global"],4),
        "Bütçe (TL)":round(r["budget"],2),
        "Durum":{"ok":"✅","over":"❌","under":"⚠️"}.get(r["status"],"—"),
        "Aktivite":f"{r['act_count']}" if r["has_acts"] else "—",
    } for r in flat])
    st.dataframe(rdf,use_container_width=True,hide_index=True,
                 column_config={
                     "  Öğe Tanımı":st.column_config.TextColumn(width="large"),
                     "Yerel %":st.column_config.NumberColumn(format="%.2f %%",width="small"),
                     "Genel %":st.column_config.NumberColumn(format="%.4f %%",width="small"),
                     "Bütçe (TL)":st.column_config.NumberColumn(format="%,.2f",width="medium"),
                 })

    st.markdown("---")
    st.markdown("### 📋 Aktivite Detay Tablosu")
    if not act_rows_data:
        st.info("Henüz aktivite eklenmedi.")
    else:
        h1,h2,h3=gh
        _ara_c = ["Ara Grup"] + [f"Ara Grup {i}" for i in range(2, _n_ara+1)]
        adf_cols=(["WBS Öğesi","WBS Adı","WBS Bütçe (TL)"] +
                  ["Ana Grup"] + _ara_c +
                  ["Aktivite No","Aktivite Adı","Tür",
                   "Alt Pay (%)","Genel Pay (%)","Bütçe (TL)",h1,h2,h3])
        adf=pd.DataFrame(act_rows_data)[adf_cols]
        st.dataframe(adf,use_container_width=True,hide_index=True,
                     column_config={
                         "WBS Bütçe (TL)":st.column_config.NumberColumn(format="%,.2f"),
                         "Alt Pay (%)":st.column_config.NumberColumn(format="%.4f %%"),
                         "Genel Pay (%)":st.column_config.NumberColumn(format="%.4f %%"),
                         "Bütçe (TL)":st.column_config.NumberColumn(format="%,.2f"),
                     })

        # ── Aktivite Para Doğrulaması ─────────────────────────────────
        st.markdown("#### ✅ Aktivite Bütçe Doğrulaması")
        st.caption("Her WBS öğesinin aktivite toplamının, öğenin kendi bütçesiyle eşleşip eşleşmediğini kontrol eder.")
        val_rows = []
        for nid, act_list in acts.items():
            n = get_node(nid)
            if not n: continue
            nb = global_pct(nid) * bedel
            bgt_m = compute_act_budgets(act_list, nb)
            # Sadece en üst seviyedeki aktivite/grupların toplamını al (direkt çocuklar)
            all_nos = {r["no"] for r in act_list}
            top_nos = [r["no"] for r in act_list
                       if "." not in r["no"] or
                          ".".join(r["no"].split(".")[:-1]) not in all_nos]
            top_total = sum(bgt_m.get(no,0) for no in top_nos)
            diff  = top_total - nb
            ok    = abs(diff) < 1.0
            val_rows.append({
                "WBS Öğesi":     node_path(nid),
                "Öğe Bütçesi (TL)": round(nb,2),
                "Aktivite Toplamı (TL)": round(top_total,2),
                "Fark (TL)":     round(diff,2),
                "Durum":         "✅ Eşit" if ok else f"⚠️ Fark: {diff:+,.2f} TL",
            })
        if val_rows:
            vdf = pd.DataFrame(val_rows)
            st.dataframe(vdf, hide_index=True, use_container_width=True,
                         column_config={
                             "Öğe Bütçesi (TL)":      st.column_config.NumberColumn(format="%,.2f"),
                             "Aktivite Toplamı (TL)":  st.column_config.NumberColumn(format="%,.2f"),
                             "Fark (TL)":              st.column_config.NumberColumn(format="%+,.2f"),
                         })
            n_ok  = sum(1 for r in val_rows if "Eşit" in r["Durum"])
            n_err = len(val_rows) - n_ok
            if n_err == 0:
                st.success(f"✅ Tüm {n_ok} WBS öğesinde aktivite toplamı bütçeyle eşleşiyor.")
            else:
                st.error(f"❌ {n_err} WBS öğesinde aktivite toplamı bütçeden farklı — "
                         f"pursantaj değerlerini kontrol edin.")

    st.markdown("---")
    st.markdown("### 📥 Raporu Dışa Aktar")
    dl1,dl2=st.columns(2)
    with dl1:
        st.download_button("📊 Excel Raporu (.xlsx)",
            data=build_excel_report(flat,act_rows_data,errors,warnings_list,info_list,bedel,gh),
            file_name="pursantaj_raporu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with dl2:
        st.download_button("📄 HTML → PDF Raporu",
            data=build_html_report(flat,act_rows_data,errors,warnings_list,info_list,bedel),
            file_name="pursantaj_raporu.html", mime="text/html",
            use_container_width=True)
    st.caption("💡 HTML dosyasını tarayıcıda açın → **🖨️ PDF Olarak Kaydet** → 'Hedef: PDF olarak kaydet'")

# ══════════════════════════════════════════════════════════════════════════
# TAB 4 — Filtre & Analiz
# ══════════════════════════════════════════════════════════════════════════
with tab4:
    gh    = st.session_state.grup_headers
    h1,h2,h3 = gh
    bedel = st.session_state.bedel
    acts  = st.session_state.activities

    all_act, wbs_depth, n_ara_grp = build_act_report(acts, bedel, gh)
    if not all_act:
        st.info("📭 Henüz aktivite yüklenmedi. Önce Aktiviteler sekmesinden Excel yükleyin.")
        st.stop()

    df_all = pd.DataFrame(all_act)
    # WBS seviye sütun adları
    lvl_cols = [f"L{i}" for i in range(wbs_depth)]

    st.markdown("### 🔍 Filtre Paneli")
    st.caption("Tüm seçimleri yapın, ardından **Filtrele** butonuna basın.")

    # Form: tüm filtreler tek seferde uygulanır → her tuşta yeniden hesaplanmaz
    with st.form("filtre_form"):
        fa,fb,fc,fd,fe = st.columns([2,1.5,1.5,1.5,1])
        with fa:
            search = st.text_input("🔎 Aktivite / Grup adında ara", placeholder="örn: asansör, kaba sıva")
        with fb:
            tür_opts = ["Tümü","Ana Grup","Ara Grup","Aktivite"]
            tür_sel  = st.selectbox("Tür", tür_opts)
        with fc:
            g1_vals = ["Tümü"] + sorted([x for x in df_all[h1].dropna().unique() if str(x).strip()])
            g1_sel  = st.selectbox(h1, g1_vals)
        with fd:
            g2_vals = ["Tümü"] + sorted([x for x in df_all[h2].dropna().unique() if str(x).strip()])
            g2_sel  = st.selectbox(h2, g2_vals)
        with fe:
            g3_vals = ["Tümü"] + sorted([x for x in df_all[h3].dropna().unique() if str(x).strip()])
            g3_sel  = st.selectbox(h3, g3_vals)

        # WBS seviye filtreleri (ağaca göre dinamik)
        wbs_filters = {}
        if wbs_depth > 0:
            wbs_cols = st.columns(min(wbs_depth, 4))
            for li in range(min(wbs_depth, 4)):
                col_key = f"L{li}"
                if col_key in df_all.columns:
                    vals = ["Tümü"] + sorted([x for x in df_all[col_key].dropna().unique() if str(x).strip()])
                    with wbs_cols[li]:
                        wbs_filters[col_key] = st.selectbox(f"WBS Seviye {li}", vals, key=f"wbs_l{li}")

        btn_col1, btn_col2 = st.columns([1,4])
        with btn_col1:
            submitted = st.form_submit_button("🔍 Filtrele", type="primary", use_container_width=True)
        with btn_col2:
            reset = st.form_submit_button("🔄 Temizle", use_container_width=False)

    # Filtre state — sadece submit sonrası değişir
    if "filt_state" not in st.session_state:
        st.session_state.filt_state = {"search":"","tür":"Tümü","g1":"Tümü","g2":"Tümü","g3":"Tümü","wbs":{}}
    if submitted:
        st.session_state.filt_state = {"search":search,"tür":tür_sel,"g1":g1_sel,"g2":g2_sel,"g3":g3_sel,"wbs":wbs_filters}
    if reset:
        st.session_state.filt_state = {"search":"","tür":"Tümü","g1":"Tümü","g2":"Tümü","g3":"Tümü","wbs":{}}

    fs = st.session_state.filt_state
    dff = df_all.copy()
    if fs["search"]:
        mask = (dff["Aktivite Adı"].str.contains(fs["search"], case=False, na=False) |
                dff["WBS Adı"].str.contains(fs["search"], case=False, na=False))
        dff = dff[mask]
    if fs["tür"] != "Tümü":
        dff = dff[dff["Tür"] == fs["tür"]]
    if fs["g1"] != "Tümü":
        dff = dff[dff[h1] == fs["g1"]]
    if fs["g2"] != "Tümü":
        dff = dff[dff[h2] == fs["g2"]]
    if fs["g3"] != "Tümü":
        dff = dff[dff[h3] == fs["g3"]]
    for lk,lv in fs.get("wbs",{}).items():
        if lv != "Tümü" and lk in dff.columns:
            dff = dff[dff[lk] == lv]

    # ── Özet metrikler ────────────────────────────────────────────────
    dff_only_act = dff[dff["Tür"]=="Aktivite"]
    dff_grp      = dff[dff["Tür"]!="Aktivite"]
    act_total    = dff_only_act["Bütçe (TL)"].sum()
    flt_total    = dff["Bütçe (TL)"].sum()
    show_total   = act_total if not dff_only_act.empty else flt_total
    show_lbl     = "Aktivite Bütçe Toplamı" if not dff_only_act.empty else "Seçili Satır Toplamı"
    show_sub     = (f"{len(dff_only_act)} aktivite satırı"
                    if not dff_only_act.empty
                    else f"{len(dff_grp)} grup satırı (çift sayım içerebilir)")
    pct_of_total = (show_total/bedel*100) if bedel>0 else 0

    st.markdown(f"""
    <div class="mrow">
      <div class="mbox"><div class="lbl">Filtrelenmiş Satır</div><div class="val">{len(dff)}</div>
        <div class="sub">{len(dff_grp)} grup + {len(dff_only_act)} aktivite</div></div>
      <div class="mbox"><div class="lbl">{show_lbl}</div><div class="val">{show_total:,.2f} TL</div>
        <div class="sub">%{pct_of_total:.4f} &nbsp;·&nbsp; {show_sub}</div></div>
      <div class="mbox"><div class="lbl">Proje Toplam</div><div class="val">{bedel:,.2f} TL</div>
        <div class="sub">Kapsama: %{pct_of_total:.2f}</div></div>
    </div>""", unsafe_allow_html=True)

    # ── Filtrelenmiş Sonuçlar ─────────────────────────────────────────
    st.markdown("---")
    total_rows = len(dff)
    PAGE = 500
    show_all = st.checkbox(f"Tümünü göster ({total_rows} satır)", value=total_rows<=PAGE, key="show_all_rows")
    dff_show = dff if show_all else dff.head(PAGE)
    if not show_all and total_rows > PAGE:
        st.caption(f"İlk {PAGE} satır gösteriliyor. Excel'e indirerek tümüne ulaşabilirsiniz.")
    st.markdown(f"#### 📋 Filtrelenmiş Sonuçlar ({total_rows} satır)")

    # Görüntüleme sütunları — WBS seviyeler + ana sütunlar
    # Dinamik Ara Grup sütunları
    ara_cols = ["Ara Grup"] + [f"Ara Grup {i}" for i in range(2, n_ara_grp+1)]
    base_show = (["WBS Öğesi","WBS Adı"] + lvl_cols +
                 ["Ana Grup"] + ara_cols +
                 ["Aktivite No","Aktivite Adı","Tür",
                  "Alt Pay (%)","Genel Pay (%)","Bütçe (TL)",h1,h2,h3])
    # Mevcut sütunlardan sadece var olanlar
    show_cols = [c for c in base_show if c in dff.columns]

    dff_display = dff_show[show_cols].copy()
    dff_display["Aktivite No"] = dff_display["Aktivite No"].str.lstrip("'")
    st.dataframe(dff_display, hide_index=True, use_container_width=True,
                 column_config={
                     "Aktivite No":  st.column_config.TextColumn(),
                     "Alt Pay (%)":  st.column_config.NumberColumn(format="%.4f %%"),
                     "Genel Pay (%)":st.column_config.NumberColumn(format="%.4f %%"),
                     "Bütçe (TL)":   st.column_config.NumberColumn(format="%,.2f"),
                     "WBS Bütçe (TL)":st.column_config.NumberColumn(format="%,.2f"),
                 })

    # ── Excel İndir ───────────────────────────────────────────────────
    def export_filtered_excel(df_src, cols):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        buf = io.BytesIO()
        wb  = __import__("openpyxl").Workbook()
        ws  = wb.active; ws.title="Filtre Sonuçları"
        ws.sheet_view.showGridLines=False
        thin=Side(style="thin",color="B0C4DE")
        def brd(): return Border(top=thin,left=thin,right=thin,bottom=thin)
        H_BG="2C5F8A"; H_FG="FFFFFF"; TEXT="1A1A2E"
        kind_bg={"Ana Grup":"D6E4F0","Ara Grup":"EAF4FB","Aktivite":"F8FBFF"}

        for ci,h_col in enumerate(cols,1):
            c=ws.cell(1,ci,h_col)
            c.fill=PatternFill("solid",fgColor=H_BG)
            c.font=Font(color=H_FG,bold=True,size=9,name="Calibri")
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=brd()
        ws.row_dimensions[1].height=18

        # TÜMÜNÜ yaz (sayfa limiti yok)
        for ri,row in enumerate(df_src[cols].itertuples(index=False),2):
            tür_val=str(getattr(row,"Tür","") if hasattr(row,"Tür") else "")
            bg=kind_bg.get(tür_val,"F8FBFF")
            for ci,(col_name,v) in enumerate(zip(cols,row),1):
                cell=ws.cell(ri,ci)
                if col_name=="Aktivite No":
                    cell.value=str(v).lstrip("'"); cell.number_format="@"
                    cell.alignment=Alignment(horizontal="left",vertical="center")
                elif isinstance(v,float) and "Bütçe" in col_name:
                    cell.value=v; cell.number_format='#,##0.00 "TL"'
                    cell.alignment=Alignment(horizontal="right",vertical="center")
                elif isinstance(v,float) and "%" in col_name:
                    cell.value=v/100 if v is not None else None
                    cell.number_format="0.0000%"
                    cell.alignment=Alignment(horizontal="right",vertical="center")
                else:
                    cell.value=v if v is not None else ""
                    cell.alignment=Alignment(horizontal="left",vertical="center")
                cell.fill=PatternFill("solid",fgColor=bg)
                cell.font=Font(color=TEXT,size=9,name="Calibri")
                cell.border=brd()
            ws.row_dimensions[ri].height=14

        for col in ws.columns:
            mx=max((len(str(c.value or "")) for c in col),default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx+3,50)
        wb.save(buf); return buf.getvalue()

    # Excel sütunları: WBS seviye sütunları + tüm veriler
    # Excel sütunları: tüm hiyerarşi sütunları dahil
    excel_cols = [c for c in base_show if c in dff.columns]
    st.download_button(
        f"📥 Tüm {total_rows} Satırı İndir (.xlsx)",
        data=export_filtered_excel(dff, excel_cols),
        file_name="filtre_sonuc.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("💡 Excel'de L0/L1/L2... sütunlarını pivot tablo kırılımı olarak kullanabilirsiniz.")


# ══════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════
# TAB 5 — Grafik: Maliyet Kırılımı Treemap
# ══════════════════════════════════════════════════════════════════════════

# ── Cached veri hazırlayıcıları (figure dışında) ──────────────────────────
@st.cache_data(show_spinner=False)
def get_wbs_treemap_data(tj: str, bedel: float):
    """WBS hiyerarşisini treemap için hazırlar. Ağaç veya bedel değişince yeniden hesaplanır."""
    tree   = json.loads(tj)
    id_map = {n["id"]: n for n in tree}

    def ch(nid):     return [n for n in tree if n["parent"] == nid]
    def gp(nid):
        n = id_map.get(nid)
        if not n: return 0.0
        if n["parent"] is None: return n["pct"] / 100
        sibs = [x for x in tree if x["parent"] == n["parent"]]
        s    = sum(x["pct"] for x in sibs)
        return gp(n["parent"]) * (n["pct"] / s if s else 0)
    def full_path(nid):
        parts, cur = [], nid
        while cur is not None:
            n = id_map.get(cur)
            if not n: break
            parts.append(n["name"]); cur = n["parent"]
        return " › ".join(reversed(parts))
    def parent_path(nid):
        n = id_map.get(nid)
        if not n or n["parent"] is None: return ""
        return full_path(n["parent"])

    rows = []
    def walk(nid, depth=0):
        n = id_map.get(nid)
        if not n: return
        b = round(gp(nid) * bedel, 2)
        rows.append({
            "path":        full_path(nid),
            "parent_path": parent_path(nid),
            "name":        n["name"],
            "depth":       depth,
            "budget":      b,
            "pct":         round(gp(nid) * 100, 4),
        })
        for k in ch(nid): walk(k["id"], depth + 1)
    for r in [n for n in tree if n["parent"] is None]:
        walk(r["id"])
    return rows


@st.cache_data(show_spinner=False)
def get_act_treemap_data(acts_json: str, bedel: float, gh_json: str):
    """Aktivite verisini Ana Grup → Ara Grup → Aktivite hiyerarşisinde hazırlar."""
    acts = json.loads(acts_json)
    gh   = json.loads(gh_json)
    h1,h2,h3 = (gh + ["Grup 1","Grup 2","Grup 3"])[:3]

    rows = []
    for nid_str, act_list in acts.items():
        nid = int(nid_str)
        n   = get_node(nid)
        if not n: continue
        nb      = global_pct(nid) * bedel
        bgt_map = compute_act_budgets(act_list, nb)
        wbs_path = node_path(nid)

        for r in act_list:
            no      = r["no"]
            bgt     = bgt_map.get(no, 0.0)
            kind    = r.get("kind","aktivite")
            # Hiyerarşi: wbs_path > no parçaları
            parts   = no.split(".")
            # path = wbs_path + > her basamak
            path    = wbs_path + " › " + no
            # parent: bir üst no veya wbs_path kendisi
            if len(parts) > 1:
                parent = wbs_path + " › " + ".".join(parts[:-1])
            else:
                parent = wbs_path
            rows.append({
                "path":        path,
                "parent_path": parent,
                "name":        r["name"],
                "no":          no,
                "depth":       len(parts),
                "budget":      round(bgt, 2),
                "pct":         round(r.get("d4",0)*100, 4) if r.get("d4") else
                               round((r.get("d3",0) or 0)*100, 4),
                "kind":        kind,
                "wbs":         n["name"],
                "wbs_path":    wbs_path,
                h1:            r.get("grup1",""),
                h2:            r.get("grup2",""),
                h3:            r.get("grup3",""),
            })
        # WBS kök node'unu da ekle (parent boş = treemap root)
        rows.append({
            "path": wbs_path, "parent_path": "",
            "name": n["name"], "no": "", "depth": 0,
            "budget": round(nb,2), "pct": round(global_pct(nid)*100,4),
            "kind": "wbs", "wbs": n["name"], "wbs_path": wbs_path,
            h1:"",h2:"",h3:"",
        })
    return rows


@st.cache_data(show_spinner=False)
def build_treemap_figure(
    rows_json: str, depth: int, metric: str,
    colorscheme: str, show_pct: bool, show_tl: bool,
    height: int, mode: str
):
    """Treemap figürünü önbellekte tutar — yalnızca parametre değişince yeniden çizer."""
    import colorsys
    rows = json.loads(rows_json)
    df   = pd.DataFrame(rows)
    if df.empty: return None

    # Derinlik filtresi
    df = df[df["depth"] <= depth].copy()
    if df.empty: return None

    val_col = "budget" if metric == "Bütçe (TL)" else "pct"
    df["val"] = df[val_col].clip(lower=0.001)

    # Renk paleti
    cs_map = {
        "Mavi (Kurumsal)":  [[0,"#0a1929"],[0.25,"#0d3b6e"],[0.6,"#1565c0"],[1,"#58a6ff"]],
        "Yeşil (Doğal)":    [[0,"#051a05"],[0.25,"#1b4a1b"],[0.6,"#2ea043"],[1,"#56d364"]],
        "Turuncu (Sıcak)":  [[0,"#1a0800"],[0.25,"#7c2d00"],[0.6,"#c84b00"],[1,"#ffa657"]],
        "Mor (Analitik)":   [[0,"#0e0a1a"],[0.25,"#3b1f6e"],[0.6,"#7c3aed"],[1,"#c084fc"]],
        "Çok Renkli":       "RdYlBu",
    }
    cscale = cs_map.get(colorscheme, cs_map["Mavi (Kurumsal)"])

    # Etiket formatı
    def fmt(r):
        # Ad: çok uzunsa kes
        name = r["name"]
        if len(name) > 28:
            name = name[:26] + "…"
        parts = [f"<b>{name}</b>"]
        if show_tl  and r["budget"] > 0:
            tl = r["budget"]
            if tl >= 1_000_000:
                parts.append(f"{tl/1_000_000:.2f}M TL")
            elif tl >= 1_000:
                parts.append(f"{tl/1_000:.1f}K TL")
            else:
                parts.append(f"{tl:,.0f} TL")
        if show_pct and r["pct"] > 0: parts.append(f"%{r['pct']:.2f}")
        return "<br>".join(parts)
    df["label_text"] = df.apply(fmt, axis=1)

    fig = go.Figure(go.Treemap(
        ids           = df["path"],
        labels        = df["name"],
        parents       = df["parent_path"],
        values        = df["val"],
        customdata    = df[["budget","pct","name"]].values,
        text          = df["label_text"],
        textinfo      = "text",
        hovertemplate = (
            "<b>%{customdata[2]}</b><br>"
            "%{customdata[0]:,.2f} TL<br>"
            "Pay: %{customdata[1]:.4f}%<extra></extra>"
        ),
        tiling  = dict(packing="squarify", pad=4),
        pathbar = dict(
            visible  = True,
            thickness= 24,
            side     = "top",
            textfont = dict(size=12, color="#c9d1d9", family="Inter"),
        ),
        marker = dict(
            colorscale  = cscale,
            colors      = df["val"].tolist(),
            showscale   = True,
            reversescale= False,
            colorbar    = dict(
                title       = dict(text=metric, font=dict(color="#8b949e",size=10)),
                tickfont    = dict(color="#8b949e",size=9),
                thickness   = 12,
                len         = 0.7,
                bgcolor     = "#161b22",
                bordercolor = "#30363d",
                borderwidth = 1,
                tickformat  = ",.0f" if metric=="Bütçe (TL)" else ".2f",
            ),
            line = dict(width=2, color="#0d1117"),
        ),
        textfont        = dict(family="Inter, Arial", size=12, color="#ffffff"),
        insidetextfont  = dict(family="Inter, Arial", size=11, color="#ffffff"),
        outsidetextfont = dict(family="Inter, Arial", size=9,  color="#8b949e"),
    ))

    fig.update_layout(
        paper_bgcolor = "#0d1117",
        plot_bgcolor  = "#0d1117",
        margin        = dict(l=0, r=10, t=6, b=0),
        height        = height,
        font          = dict(family="Inter, Arial", color="#c9d1d9"),
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════

with tab5:
    import difflib

    bedel = st.session_state.bedel
    gh    = st.session_state.grup_headers
    acts  = st.session_state.activities
    tj    = json.dumps(st.session_state.tree, sort_keys=True)

    # ── Veri yükle (cached) ───────────────────────────────────────────
    wbs_rows  = get_wbs_treemap_data(tj, bedel)
    df_wbs_t5 = pd.DataFrame(wbs_rows)
    max_d_wbs = int(df_wbs_t5["depth"].max()) if not df_wbs_t5.empty else 2

    has_acts = bool(acts)
    if has_acts:
        acts_json = json.dumps({str(k):v for k,v in acts.items()}, ensure_ascii=False, sort_keys=True)
        gh_json   = json.dumps(gh, ensure_ascii=False)
        act_rows  = get_act_treemap_data(acts_json, bedel, gh_json)
        df_act_t5 = pd.DataFrame(act_rows)
        max_d_act = int(df_act_t5["depth"].max()) if not df_act_t5.empty else 3
    else:
        act_rows  = []; df_act_t5 = pd.DataFrame(); max_d_act = 3

    # ── Kontrol paneli ─────────────────────────────────────────────────
    ctrl, chart_area = st.columns([1, 5])

    with ctrl:
        st.markdown("##### ⚙️ Ayarlar")

        mode = st.radio(
            "Görünüm",
            ["🏗️ WBS Seviyesi", "📋 Aktivite Türü", "🔀 Konsolide"],
            key="t5_mode",
            help="WBS: ağaç | Aktivite: her WBS kendi aktivitelerini gösterir | Konsolide: tüm projedeki aynı kalemleri toplar"
        )
        is_wbs  = mode == "🏗️ WBS Seviyesi"
        is_act  = mode == "📋 Aktivite Türü"
        is_cons = mode == "🔀 Konsolide"

        if (is_act or is_cons) and not has_acts:
            st.warning("Aktivite yüklü değil.")

        max_d = max_d_wbs if is_wbs else max_d_act
        depth = st.slider("Derinlik", 1, max(max_d,4), min(3,max_d), key="t5_depth")

        metric      = st.radio("Metrik", ["Bütçe (TL)","Ağırlık (%)"], key="t5_metric")
        colorscheme = st.selectbox("Renk",
            ["Mavi (Kurumsal)","Yeşil (Doğal)","Turuncu (Sıcak)","Mor (Analitik)","Çok Renkli"],
            key="t5_color")
        show_tl  = st.checkbox("TL göster", value=True,  key="t5_tl")
        show_pct = st.checkbox("% göster",  value=True,  key="t5_pct")
        st.markdown("---")
        st.markdown(f"Sözleşme: **{bedel:,.0f} TL**")

    # ═══════════════════════════════════════════════════════════════════
    # Konsolide mod: fuzzy eşleştirme + onay akışı
    # ═══════════════════════════════════════════════════════════════════
    def normalize(s: str) -> str:
        """Türkçe büyük/küçük harf güvenli normalize."""
        tr_map = str.maketrans("çğışöüÇĞİŞÖÜ", "ÇĞİŞÖÜÇĞİŞÖÜ")
        return s.strip().upper().translate(tr_map)

    @st.cache_data(show_spinner=False)
    def find_fuzzy_groups(acts_json: str, bedel: float, gh_json: str, threshold: float = 0.90):
        """
        Tüm aktivite adlarını normalize eder, %threshold üzeri benzerleri gruplar.
        Returns: {canonical_name: [original_name1, original_name2, ...]}
        """
        _, _depth_unused = (lambda r: (r[0],r[1]))(build_act_report(
            json.loads(acts_json), bedel, json.loads(gh_json)
        )) if False else ([], 0)  # type hint dummy

        act_data = json.loads(acts_json)
        all_names = []
        for act_list in act_data.values():
            if not isinstance(act_list, list): continue
            for r in act_list:
                name = r.get("name","").strip()
                if name and r.get("kind") == "aktivite":
                    all_names.append(name)
        unique_names = list(dict.fromkeys(all_names))  # preserve first-seen order

        # Union-Find yapısı
        parent = {n: n for n in unique_names}
        def find(x):
            while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
            return x
        def union(a, b):
            ra, rb = find(a), find(b)
            if ra != rb: parent[rb] = ra

        norms = {n: normalize(n) for n in unique_names}
        norm_list = [(n, norms[n]) for n in unique_names]

        merges = []
        for i in range(len(norm_list)):
            for j in range(i+1, len(norm_list)):
                a, na = norm_list[i]; b, nb = norm_list[j]
                ratio = difflib.SequenceMatcher(None, na, nb).ratio()
                if ratio >= threshold:
                    merges.append((a, b, ratio))
                    union(a, b)

        # Grupla: canonical = en uzun ortak normalize isim (ilk karşılaşılan)
        groups: dict[str, list] = {}
        for n in unique_names:
            root = find(n)
            groups.setdefault(root, []).append(n)

        # Sadece 1'den fazla üyeye sahip gruplar = birleştirmeler
        merged = {k: sorted(v, key=len, reverse=True) for k,v in groups.items() if len(v)>1}
        return merged, merges

    def build_consolidated_rows(acts_dict: dict, bedel: float, rename_map: dict) -> list:
        """
        Hiyerarşik konsolidasyon — çift sayım YOK:
        - Yalnızca kind=="aktivite" yaprak satırları bütçeye eklenir
        - Tüm üst düğümler bottom-up ile yapraklardan hesaplanır
        - Pct = TL / bedel × 100  (ham d3/d4 toplanmaz → %288 gibi anlamsız değer oluşmaz)
        """
        ROOT = "Proje Geneli (Konsolide)"
        path_data: dict[str, dict] = {
            ROOT: {"name": ROOT, "parent_path": "", "depth": 0, "budget_sum": 0.0}
        }

        for nid_str, act_list in acts_dict.items():
            if not isinstance(act_list, list): continue
            nid = int(nid_str)
            n   = get_node(nid)
            if not n: continue
            nb      = round(global_pct(nid) * bedel, 2)
            bgt_map = compute_act_budgets(act_list, nb)
            no_map  = {r.get("no",""): r for r in act_list}

            for r in act_list:
                # ── Sadece gerçek yaprak (aktivite) işle ─────────────
                if r.get("kind") != "aktivite":
                    continue
                no     = r.get("no","")
                budget = bgt_map.get(no, 0.0)
                parts  = no.split(".")

                # Her ata seviyesi için path oluştur (bütçe sadece yaprağa eklenir)
                for depth_i in range(1, len(parts) + 1):
                    anc_no  = ".".join(parts[:depth_i])
                    anc_row = no_map.get(anc_no)
                    if not anc_row: continue

                    # Canonical isim zinciri
                    canonical_chain = [ROOT]
                    for d in range(1, depth_i + 1):
                        pno  = ".".join(parts[:d])
                        prow = no_map.get(pno)
                        if prow:
                            cname = rename_map.get(prow.get("name",""), prow.get("name",""))
                            if cname.strip():
                                canonical_chain.append(cname)

                    path     = " › ".join(canonical_chain)
                    parent_p = " › ".join(canonical_chain[:-1])
                    name     = canonical_chain[-1]

                    if path not in path_data:
                        path_data[path] = {
                            "name":        name,
                            "parent_path": parent_p,
                            "depth":       depth_i,
                            "budget_sum":  0.0,
                        }

                    # Bütçeyi YALNIZCA en derin seviyede (yaprak) ekle
                    if depth_i == len(parts):
                        path_data[path]["budget_sum"] += budget

        # ── Bottom-up: üstler yapraklardan hesaplanır ─────────────────
        for p in sorted(path_data, key=lambda x: path_data[x]["depth"], reverse=True):
            pp = path_data[p]["parent_path"]
            if pp and pp in path_data:
                path_data[pp]["budget_sum"] += path_data[p]["budget_sum"]

        # ── Rows ──────────────────────────────────────────────────────
        rows = []
        for path, d in path_data.items():
            budget = round(d["budget_sum"], 2)
            pct    = round(budget / bedel * 100, 4) if bedel > 0 else 0.0
            rows.append({
                "path":        path,
                "parent_path": d["parent_path"],
                "name":        d["name"],
                "depth":       d["depth"],
                "budget":      budget,
                "pct":         pct,
            })
        return rows

    # ═══════════════════════════════════════════════════════════════════
    # Chart area
    # ═══════════════════════════════════════════════════════════════════
    with chart_area:

        # ── KONSOLIDE MOD ─────────────────────────────────────────────
        if is_cons:
            if not has_acts:
                st.info("Önce Aktiviteler sekmesinden Excel yükleyin.")
            else:
                # Eşleştirme bölümü
                st.markdown("### 🔀 Konsolide Görünüm")

                merged_groups, all_merges = find_fuzzy_groups(acts_json, bedel, gh_json, 0.90)

                if "cons_rename_map" not in st.session_state:
                    st.session_state.cons_rename_map = {}
                if "cons_confirmed" not in st.session_state:
                    st.session_state.cons_confirmed = False
                if "cons_show_panel" not in st.session_state:
                    st.session_state.cons_show_panel = not st.session_state.cons_confirmed

                # Önceki kararlar — her zaman özet göster
                prev_renames = {k:v for k,v in st.session_state.cons_rename_map.items() if k!=v}
                decisions_exist = bool(st.session_state.get("cons_decisions"))

                if st.session_state.cons_confirmed:
                    if prev_renames:
                        st.success(
                            f"✅ **{len(prev_renames)} eşleştirme aktif** — "
                            + ", ".join(f'"{k}" → "{v}"' for k,v in list(prev_renames.items())[:3])
                            + (f" ve {len(prev_renames)-3} tane daha." if len(prev_renames)>3 else ".")
                        )
                    else:
                        st.info("✅ Eşleştirme onaylandı — değişiklik yapılmadı.")

                    # Mevcut kararları tablo olarak göster
                    if decisions_exist:
                        with st.expander("📋 Mevcut Eşleştirme Kararlarını Görüntüle", expanded=False):
                            dec_rows = []
                            for vk, dec in st.session_state.cons_decisions.items():
                                v = dec.get("variant","")
                                a = dec.get("action","Ayrı tut")
                                result = (dec.get("target","") if a=="Birleştir →"
                                         else dec.get("new_name","") if a=="Yeniden adlandır"
                                         else v)
                                changed = (result != v)
                                dec_rows.append({
                                    "Orijinal Ad":   v,
                                    "Karar":         a,
                                    "Sonuç":         result,
                                    "Değişti mi":    "✅ Evet" if changed else "➖ Hayır",
                                })
                            st.dataframe(pd.DataFrame(dec_rows), hide_index=True, use_container_width=True)

                # Buton: paneli aç/kapat
                btn_lbl = "✏️ Eşleştirmeleri Düzenle" if st.session_state.cons_confirmed else "🔍 Eşleştirme Yap"
                if st.button(btn_lbl, key="cons_open_btn"):
                    st.session_state.cons_show_panel = True
                    st.rerun()

                # Birleştirme önerileri — kullanıcı her varyantı bağımsız yönetir
                if merged_groups and st.session_state.cons_show_panel:
                    with st.expander(
                        f"🔍 {len(merged_groups)} benzer isim grubu bulundu — gözden geçirin",
                        expanded=True
                    ):
                        st.caption(
                            "Sistem %90+ benzerlik gösteren isimleri tespit etti. "
                            "Her varyant için **ne yapılacağını siz seçin**: "
                            "birleştir, ayrı tut veya farklı bir ad ver."
                        )
                        st.markdown("---")

                        # Geçici kararlar — form dışında tutuyoruz, form sadece submit için
                        if "cons_decisions" not in st.session_state:
                            st.session_state.cons_decisions = {}

                        decisions = st.session_state.cons_decisions

                        for gi, (root, variants) in enumerate(merged_groups.items()):
                            group_key = f"g{gi}"

                            # Benzerlik hesapla (en düşük çifti göster)
                            sims = []
                            for i in range(len(variants)):
                                for j in range(i+1, len(variants)):
                                    s = difflib.SequenceMatcher(
                                        None, normalize(variants[i]), normalize(variants[j])
                                    ).ratio()
                                    sims.append(s)
                            min_sim = min(sims) if sims else 1.0
                            max_sim = max(sims) if sims else 1.0

                            # Renk: yüksek benzerlik = sarı uyarı, düşük = kırmızı dikkat
                            sim_icon = "🟡" if max_sim >= 0.95 else "🟠" if max_sim >= 0.90 else "🔴"

                            st.markdown(
                                f"{sim_icon} **Grup {gi+1}** — "
                                f"benzerlik %{min_sim*100:.0f}–%{max_sim*100:.0f}"
                            )

                            # Her varyant için ayrı karar satırı
                            for vi, variant in enumerate(variants):
                                var_key = f"{group_key}_v{vi}"
                                col_name, col_action, col_custom = st.columns([2, 1.2, 2])

                                with col_name:
                                    st.markdown(f"`{variant}`")

                                # Varsayılan karar: ayrı tut
                                default_action = decisions.get(var_key, {}).get("action", "Ayrı tut")
                                with col_action:
                                    action = st.selectbox(
                                        "Karar",
                                        ["Ayrı tut", "Birleştir →", "Yeniden adlandır"],
                                        index=["Ayrı tut","Birleştir →","Yeniden adlandır"].index(default_action),
                                        key=f"act_{var_key}",
                                        label_visibility="collapsed",
                                    )

                                with col_custom:
                                    if action == "Birleştir →":
                                        # Hangi ada birleşsin?
                                        target_opts = [v for v in variants if v != variant]
                                        prev_target = decisions.get(var_key, {}).get("target", target_opts[0] if target_opts else variant)
                                        if prev_target not in target_opts: prev_target = target_opts[0] if target_opts else variant
                                        target = st.selectbox(
                                            "Hedef",
                                            target_opts,
                                            index=target_opts.index(prev_target) if prev_target in target_opts else 0,
                                            key=f"tgt_{var_key}",
                                            label_visibility="collapsed",
                                        )
                                        decisions[var_key] = {"action": action, "target": target, "variant": variant}
                                    elif action == "Yeniden adlandır":
                                        prev_name = decisions.get(var_key, {}).get("new_name", variant)
                                        new_name = st.text_input(
                                            "Yeni ad",
                                            value=prev_name,
                                            key=f"nm_{var_key}",
                                            label_visibility="collapsed",
                                            placeholder="Yeni ad girin...",
                                        )
                                        decisions[var_key] = {"action": action, "new_name": new_name, "variant": variant}
                                    else:
                                        st.markdown("<span style='color:#6e7681;font-size:.85rem'>Değişmez</span>",
                                                    unsafe_allow_html=True)
                                        decisions[var_key] = {"action": action, "variant": variant}

                            st.session_state.cons_decisions = decisions
                            st.markdown("---")

                        # Özet önizleme
                        merges_preview = []
                        renames_preview = []
                        for vk, dec in decisions.items():
                            v = dec.get("variant","")
                            if dec["action"] == "Birleştir →":
                                merges_preview.append(f'"{v}" → "{dec.get("target","")}"')
                            elif dec["action"] == "Yeniden adlandır":
                                renames_preview.append(f'"{v}" → "{dec.get("new_name","")}"')

                        if merges_preview or renames_preview:
                            summary_parts = []
                            if merges_preview:
                                summary_parts.append(f"**{len(merges_preview)} birleştirme:** " + ", ".join(merges_preview[:3]) + ("..." if len(merges_preview)>3 else ""))
                            if renames_preview:
                                summary_parts.append(f"**{len(renames_preview)} yeniden adlandırma:** " + ", ".join(renames_preview[:3]) + ("..." if len(renames_preview)>3 else ""))
                            st.info("📝 Yapılacak değişiklikler: " + " | ".join(summary_parts))
                        else:
                            st.info("ℹ️ Hiçbir değişiklik seçilmedi — tüm adlar orijinal haliyle kalacak.")

                        if st.button("✅ Kararları Onayla ve Grafiği Çiz", type="primary", key="cons_submit"):
                            # rename_map oluştur
                            final_map: dict[str, str] = {}
                            for vk, dec in decisions.items():
                                v = dec.get("variant","")
                                if dec["action"] == "Birleştir →":
                                    final_map[v] = dec.get("target", v)
                                elif dec["action"] == "Yeniden adlandır":
                                    nn = dec.get("new_name","").strip()
                                    final_map[v] = nn if nn else v
                                # "Ayrı tut" → rename_map'e eklenmez, kendi adını korur

                            st.session_state.cons_rename_map = final_map
                            st.session_state.cons_confirmed  = True
                            st.session_state.cons_show_panel = False  # paneli kapat

                            # Özet mesaj
                            merged_c = [(k,v) for k,v in final_map.items() if k!=v]
                            if merged_c:
                                st.success(
                                    f"✅ Onaylandı — {len(merged_c)} değişiklik uygulandı: "
                                    + ", ".join(f'"{k}" → "{v}"' for k,v in merged_c[:4])
                                    + (f" ve {len(merged_c)-4} tane daha." if len(merged_c)>4 else ".")
                                )
                            else:
                                st.success("✅ Onaylandı — tüm adlar orijinal haliyle korundu.")
                            st.rerun()
                else:
                    # Hiç uyumsuzluk yok
                    if not st.session_state.cons_confirmed:
                        st.success("✅ Tüm aktivite adları uyumlu — uyumsuzluk bulunamadı.")
                        st.session_state.cons_rename_map = {}
                        st.session_state.cons_confirmed  = True

                # Sıfırla butonu
                if st.session_state.cons_confirmed:
                    if st.button("🔄 Eşleştirmeyi Sıfırla", key="cons_reset"):
                        st.session_state.cons_confirmed   = False
                        st.session_state.cons_rename_map  = {}
                        st.session_state.cons_decisions   = {}
                        st.session_state.cons_show_panel  = True
                        st.rerun()

                # Grafik — sadece onay sonrası
                if st.session_state.cons_confirmed:
                    rename_map = st.session_state.cons_rename_map
                    # Tüm aktivite adlarını rename_map üzerinden geçir
                    full_rename = {}
                    for act_list in acts.values():
                        if not isinstance(act_list, list): continue
                        for r in act_list:
                            nm = r.get("name","")
                            full_rename[nm] = rename_map.get(nm, nm)

                    cons_rows = build_consolidated_rows(
                        {str(k):v for k,v in acts.items()}, bedel, full_rename
                    )

                    with ctrl:
                        st.markdown("---")
                        unique_acts = len(set(
                            r["name"] for r in cons_rows if r["depth"]==2
                        ))
                        st.markdown(f"**{unique_acts}** benzersiz aktivite")
                        total_c = sum(r["budget"] for r in cons_rows if r["depth"]==2)
                        st.markdown(f"**{total_c:,.0f} TL**")

                    rows_json = json.dumps(cons_rows, ensure_ascii=False, sort_keys=True)
                    fig = build_treemap_figure(
                        rows_json=rows_json, depth=min(depth,2),
                        metric=metric, colorscheme=colorscheme,
                        show_pct=show_pct, show_tl=show_tl,
                        height=max(540,680), mode=mode,
                    )
                    if fig:
                        st.plotly_chart(fig, use_container_width=True, config={
                            "displayModeBar":True,
                            "modeBarButtonsToRemove":["lasso2d","select2d"],
                            "displaylogo":False,
                            "toImageButtonOptions":{"format":"png","filename":"konsolide","height":900,"width":1600,"scale":2},
                        })

                    with st.expander("📋 Konsolide Veri Tablosu", expanded=False):
                        df_c = pd.DataFrame([
                            r for r in cons_rows if r["depth"] > 0
                        ])[["name","budget","pct"]].rename(columns={
                            "name":"Aktivite/Grup","budget":"Toplam Bütçe (TL)","pct":"Toplam Ağırlık (%)"
                        }).sort_values("Toplam Bütçe (TL)", ascending=False)
                        st.dataframe(df_c, hide_index=True, use_container_width=True,
                                     column_config={
                                         "Toplam Bütçe (TL)": st.column_config.NumberColumn(format="%,.2f"),
                                         "Toplam Ağırlık (%)":st.column_config.NumberColumn(format="%.4f %%"),
                                     })

        # ── WBS veya AKTİVİTE MODU ────────────────────────────────────
        else:
            rows_to_use = wbs_rows if is_wbs else act_rows
            if not rows_to_use:
                st.info("Veri yok. WBS ağacı veya aktivite yükleyin.")
            else:
                with ctrl:
                    st.markdown("---")
                    shown_n = len([r for r in rows_to_use if r.get("depth",0) <= depth])
                    st.markdown(f"**{shown_n}** öğe")

                chart_h   = max(540, min(max_d * 130, 680))
                rows_json = json.dumps(rows_to_use, ensure_ascii=False, sort_keys=True)
                fig = build_treemap_figure(
                    rows_json=rows_json, depth=depth, metric=metric,
                    colorscheme=colorscheme, show_pct=show_pct, show_tl=show_tl,
                    height=chart_h, mode=mode,
                )
                if fig:
                    st.plotly_chart(fig, use_container_width=True, config={
                        "displayModeBar":True,
                        "modeBarButtonsToRemove":["lasso2d","select2d"],
                        "displaylogo":False,
                        "toImageButtonOptions":{"format":"png","filename":"maliyet_kirilimi","height":900,"width":1600,"scale":2},
                    })
                else:
                    st.info("Seçili derinlikte gösterilecek veri yok.")

                with st.expander("📋 Veri Tablosu", expanded=False):
                    df_show = pd.DataFrame(rows_to_use)
                    df_show = df_show[df_show["depth"] <= depth][["depth","name","budget","pct"]]
                    df_show.columns = ["Seviye","Öğe","Bütçe (TL)","Ağırlık (%)"]
                    df_show = df_show.sort_values(["Seviye","Bütçe (TL)"], ascending=[True,False])
                    st.dataframe(df_show, hide_index=True, use_container_width=True,
                                 column_config={
                                     "Bütçe (TL)": st.column_config.NumberColumn(format="%,.2f"),
                                     "Ağırlık (%)":st.column_config.NumberColumn(format="%.4f %%"),
                                 })


