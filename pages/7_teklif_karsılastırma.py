"""
Teklif Kıyaslama ve Karar Destek Sistemi (Nihai & Stabil Sürüm)
===============================================================
- Excel Şablonuna "Açık Sarı" özel veri giriş alanları (Firma, Tarih) eklendi.
- Excel Şablonuna "Tarih" ve "Sayısal Veri" Data Validation (Hata Koruması) eklendi.
- Taşeronun metin/format bozması engellendi.
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
import io
import plotly.graph_objects as go
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# st.set_page_config(page_title="Dinamik Teklif Kıyaslama", page_icon="📊", layout="wide")

# ==========================================
# 1. YARDIMCI FONKSİYONLAR VE FORMATLAYICILAR
# ==========================================
def tr_format(val):
    if pd.isna(val) or val == "": return "0,00"
    try:
        formatted = "{:,.2f}".format(float(val))
        return formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return val

def safe_float(val):
    if pd.isna(val) or val == "": return 0.0
    try:
        val_str = str(val).strip()
        if ',' in val_str and '.' in val_str:
            val_str = val_str.replace('.', '').replace(',', '.')
        elif ',' in val_str:
            val_str = val_str.replace(',', '.')
        return float(val_str)
    except:
        return 0.0

# ==========================================
# 2. VERİ YAPISI VE HAFIZA (SESSION STATE)
# ==========================================
def init_state():
    if "proje_adi" not in st.session_state: st.session_state.proje_adi = ""
    if "is_kalemleri" not in st.session_state: 
        st.session_state.is_kalemleri = pd.DataFrame(columns=["İş Grubu", "İmalat Adı", "Açıklama", "Özellikler", "Birim", "Metraj"])
    if "firmalar" not in st.session_state: st.session_state.firmalar = {} 
    if "teklifler" not in st.session_state: st.session_state.teklifler = {}

init_state()

def export_project_to_json():
    df_kalemler = st.session_state.get("is_kalemleri_guncel", st.session_state.is_kalemleri)
    data = {
        "proje_adi": st.session_state.proje_adi,
        "is_kalemleri": df_kalemler.to_dict(orient="records"),
        "firmalar": st.session_state.firmalar,
        "teklifler": {f: df.to_dict(orient="records") for f, df in st.session_state.teklifler.items()}
    }
    return json.dumps(data, indent=4, ensure_ascii=False)

def import_project_from_json(json_file):
    try:
        data = json.load(json_file)
        st.session_state.proje_adi = data.get("proje_adi", "")
        df_kalemler = pd.DataFrame(data.get("is_kalemleri", []))
        if "Metraj" in df_kalemler.columns: df_kalemler["Metraj"] = df_kalemler["Metraj"].apply(safe_float)
        st.session_state.is_kalemleri = df_kalemler
        
        yuklenen_firmalar = data.get("firmalar", {})
        if isinstance(yuklenen_firmalar, list): 
            st.session_state.firmalar = {f: {"yetkili": "", "telefon": "", "not": "", "teklif_tarihi": "", "gecerlilik_tarihi": ""} for f in yuklenen_firmalar}
        else: 
            st.session_state.firmalar = yuklenen_firmalar
            
        st.session_state.teklifler = {f: pd.DataFrame(df_data) for f, df_data in data.get("teklifler", {}).items()}
        return True
    except Exception:
        return False

# ==========================================
# 3. PROFESYONEL EXCEL ÇIKTI ÜRETİCİLERİ
# ==========================================
def generate_firm_excel(df_firma):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_firma.to_excel(writer, index=False, sheet_name="Firma_Teklifi")
    return output.getvalue()

def generate_empty_template(df_items, proje_adi):
    output = io.BytesIO()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Teklif_Sablonu"
    ws.page_setup.paperSize = 8; ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Uyarıcı Açık Sarı
    
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Veri Doğrulama (Data Validation) Kuralları
    dv_tarih = DataValidation(type="date", operator="greaterThanOrEqual", formula1="44000", allow_blank=True)
    dv_tarih.errorTitle = 'Geçersiz Format'
    dv_tarih.error = 'Lütfen hücreye sadece geçerli bir tarih giriniz (Örn: 01.01.2026).'
    ws.add_data_validation(dv_tarih)
    
    dv_fiyat = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_fiyat.errorTitle = 'Sayısal Hata'
    dv_fiyat.error = 'Lütfen bu hücreye sadece rakam / fiyat giriniz. Metin girmek yasaktır.'
    ws.add_data_validation(dv_fiyat)

    # 1-2. Satırlar: Proje Adı
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
    p_name = proje_adi if proje_adi else "PROJE ADI GİRİNİZ"
    proje_hucre = ws.cell(row=1, column=1, value=f"PROJE ADI: {p_name.upper()}\nALT YÜKLENİCİ FİYAT TEKLİF FORMU (Lütfen sadece sarı alanları doldurunuz)")
    proje_hucre.font = Font(bold=True, size=12); proje_hucre.alignment = center_align; proje_hucre.fill = label_fill
    for r in range(1, 3):
        for c in range(1, 9): ws.cell(row=r, column=c).border = thin_border
    
    # 3. Satır: Taşeron Bilgi Giriş Alanı (İzole ve Korumalı)
    ws.row_dimensions[3].height = 25
    
    # Firma Adı
    ws.cell(row=3, column=1, value="FİRMA ADI:").font = bold_font
    ws.cell(row=3, column=1).fill = label_fill; ws.cell(row=3, column=1).border = thin_border; ws.cell(row=3, column=1).alignment = left_align
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
    ws.cell(row=3, column=2).fill = input_fill; ws.cell(row=3, column=2).border = thin_border; ws.cell(row=3, column=3).border = thin_border
    
    # Teklif Tarihi
    ws.cell(row=3, column=4, value="TEKLİF TARİHİ:").font = bold_font
    ws.cell(row=3, column=4).fill = label_fill; ws.cell(row=3, column=4).border = thin_border; ws.cell(row=3, column=4).alignment = center_align
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6)
    c_tt = ws.cell(row=3, column=5)
    c_tt.fill = input_fill; c_tt.border = thin_border; c_tt.alignment = center_align; c_tt.number_format = 'DD.MM.YYYY'
    ws.cell(row=3, column=6).border = thin_border
    dv_tarih.add('E3:F3')
    
    # Geçerlilik Tarihi
    ws.cell(row=3, column=7, value="GEÇERLİLİK TRH:").font = bold_font
    ws.cell(row=3, column=7).fill = label_fill; ws.cell(row=3, column=7).border = thin_border; ws.cell(row=3, column=7).alignment = center_align
    c_gt = ws.cell(row=3, column=8)
    c_gt.fill = input_fill; c_gt.border = thin_border; c_gt.alignment = center_align; c_gt.number_format = 'DD.MM.YYYY'
    dv_tarih.add('H3')

    # 4. Satır: Sütun Başlıkları
    headers = ["İş Grubu", "İmalat Adı", "Açıklama", "Özellikler", "Birim", "Metraj", "Birim Fiyat", "Toplam Tutar"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
        
    current_row = 5
    if df_items is not None and not df_items.empty:
        for _, row in df_items.iterrows():
            ws.row_dimensions[current_row].height = 35 
            for i, col in enumerate(headers[:-2], 1): 
                val = row.get(col, "")
                c = ws.cell(row=current_row, column=i, value=str(val) if pd.notna(val) else "")
                c.border = thin_border; c.alignment = center_align
            
            # Birim Fiyat (Sarı Dolgu ve Sayısal Doğrulama)
            c_bf = ws.cell(row=current_row, column=7, value="")
            c_bf.fill = input_fill; c_bf.border = thin_border; c_bf.number_format = '#,##0.00'
            dv_fiyat.add(f'G{current_row}')
            
            c_toplam = ws.cell(row=current_row, column=8, value=f'=IF(ISNUMBER(G{current_row}), F{current_row}*G{current_row}, "")')
            c_toplam.border = thin_border; c_toplam.number_format = '#,##0.00'; c_toplam.font = Font(bold=True); c_toplam.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            current_row += 1
    else:
        for _ in range(15):
            ws.row_dimensions[current_row].height = 35
            for i in range(1, 8): ws.cell(row=current_row, column=i).border = thin_border
            
            ws.cell(row=current_row, column=7).fill = input_fill
            dv_fiyat.add(f'G{current_row}')
            
            ws.cell(row=current_row, column=8, value=f'=IF(ISNUMBER(G{current_row}), F{current_row}*G{current_row}, "")').border = thin_border
            current_row += 1

    ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25; ws.column_dimensions['E'].width = 10; ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15; ws.column_dimensions['H'].width = 18

    wb.save(output)
    return output.getvalue()

def generate_a3_excel_report(df_main, is_kalemleri, firmalar_dict, anonim=False):
    output = io.BytesIO()
    from openpyxl import Workbook
    if anonim:
        anonim_mapping = {}; anonim_firmalar_dict = {}
        for i, firma in enumerate(firmalar_dict.keys(), 1):
            anon_name = f"FİRMA {i}"
            anonim_mapping[firma] = anon_name
            anonim_firmalar_dict[anon_name] = {"yetkili": "Gizli", "telefon": "Gizli", "not": "Anonim Rapor", "teklif_tarihi": "Gizli", "gecerlilik_tarihi": "Gizli"}
        rename_dict = {}
        for col in df_main.columns:
            for f in firmalar_dict.keys():
                if col.startswith(f"{f} -"): rename_dict[col] = col.replace(f"{f} -", f"{anonim_mapping[f]} -")
        df_main = df_main.rename(columns=rename_dict); firmalar_dict = anonim_firmalar_dict
    
    wb = Workbook(); ws = wb.active; ws.title = "Kıyaslama"
    ws.page_setup.paperSize = 8; ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    
    firmalar_listesi = list(firmalar_dict.keys())
    firm_totals = {f: df_main[f"{f} - Toplam"].sum() if f"{f} - Toplam" in df_main.columns else 0 for f in firmalar_listesi}
    ranked_firms = sorted([(f, v) for f, v in firm_totals.items() if v > 0], key=lambda x: x[1])
    rank_dict = {f: i+1 for i, (f, v) in enumerate(ranked_firms)}
    
    optimum_total = 0
    bf_cols = [f"{f} - B.Fiyat" for f in firmalar_listesi if f"{f} - B.Fiyat" in df_main.columns]
    if bf_cols:
        df_opt_zero_nan = df_main[bf_cols].replace(0, np.nan)
        optimum_total = (df_main["Metraj"] * df_opt_zero_nan.min(axis=1).fillna(0)).sum()

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"); subtotal_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid") 
    min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"); min_font = Font(color="006100", bold=True)
    max_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"); max_font = Font(color="9C0006", bold=False)
    bold_font = Font(bold=True); center_align = Alignment(horizontal="center", vertical="center", wrap_text=True); left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=7, end_column=6)
    p_name = st.session_state.proje_adi if st.session_state.proje_adi else "PROJE ADI GİRİNİZ"
    proje_hucre = ws.cell(row=1, column=1, value=f"PROJE ADI:\n{p_name.upper()}"); proje_hucre.font = Font(bold=True, size=14); proje_hucre.alignment = center_align
    for r in range(1, 8):
        for c in range(1, 7): ws.cell(row=r, column=c).border = thin_border

    col_idx = 7
    for firma in firmalar_listesi:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
        ws.cell(row=1, column=col_idx, value=firma).font = bold_font; ws.cell(row=1, column=col_idx).fill = header_fill; ws.cell(row=1, column=col_idx).alignment = center_align
        ws.cell(row=1, column=col_idx).border = thin_border; ws.cell(row=1, column=col_idx+1).border = thin_border
        
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+1)
        ws.cell(row=2, column=col_idx, value=f"{rank_dict.get(firma, '-')} . SIRA  |  {firm_totals[firma]:,.2f} ₺").font = bold_font
        ws.cell(row=2, column=col_idx).alignment = center_align; ws.cell(row=2, column=col_idx).border = thin_border; ws.cell(row=2, column=col_idx+1).border = thin_border

        bilgiler = [
            (3, "Yetkili", "yetkili"), 
            (4, "Tel", "telefon"), 
            (5, "T. Tarihi", "teklif_tarihi"),
            (6, "Geçerlilik", "gecerlilik_tarihi"),
            (7, "Not", "not")
        ]
        for r_idx, label, key in bilgiler:
            ws.merge_cells(start_row=r_idx, start_column=col_idx, end_row=r_idx, end_column=col_idx+1)
            ws.cell(row=r_idx, column=col_idx, value=f"{label}: {firmalar_dict[firma].get(key, '')}").alignment = left_align
            ws.cell(row=r_idx, column=col_idx).border = thin_border; ws.cell(row=r_idx, column=col_idx+1).border = thin_border
        col_idx += 2
        
    headers = ["İş Grubu", "İmalat Adı", "Açıklama", "Özellikler", "Birim", "Metraj"]
    for i, h in enumerate(headers, 1): ws.cell(row=8, column=i, value=h).font = bold_font; ws.cell(row=8, column=i).alignment = center_align; ws.cell(row=8, column=i).border = thin_border
        
    col_idx = 7
    for _ in firmalar_listesi:
        for sub_h in ["Birim Fiyat", "Toplam Tutar"]:
            ws.cell(row=8, column=col_idx, value=sub_h).font = bold_font; ws.cell(row=8, column=col_idx).alignment = center_align; ws.cell(row=8, column=col_idx).border = thin_border
            col_idx += 1

    current_row = 9
    is_kalemleri["İş Grubu"] = is_kalemleri["İş Grubu"].fillna("Genel")
    df_main["İş Grubu"] = df_main["İş Grubu"].fillna("Genel")
    
    for grup in is_kalemleri["İş Grubu"].unique():
        grup_df = df_main[df_main["İş Grubu"] == grup]
        grup_toplamlari = {firma: 0.0 for firma in firmalar_listesi}
        
        for _, row in grup_df.iterrows():
            ws.row_dimensions[current_row].height = 40
            for col_i, val in enumerate(["İş Grubu", "İmalat Adı", "Açıklama", "Özellikler", "Birim"], 1):
                c = ws.cell(row=current_row, column=col_i, value=str(row.get(val, "")))
                c.border = thin_border; c.alignment = center_align 
                
            c_metraj = ws.cell(row=current_row, column=6, value=safe_float(row.get("Metraj", 0)))
            c_metraj.border = thin_border; c_metraj.alignment = center_align; c_metraj.number_format = '#,##0.00'
            
            fiyatlar = [safe_float(row.get(f"{f} - B.Fiyat", 0)) for f in firmalar_listesi if safe_float(row.get(f"{f} - B.Fiyat", 0)) > 0]
            min_fiyat = min(fiyatlar) if fiyatlar else None
            max_fiyat = max(fiyatlar) if fiyatlar else None

            c_idx = 7
            for firma in firmalar_listesi:
                bf = safe_float(row.get(f"{firma} - B.Fiyat", 0)); top = safe_float(row.get(f"{firma} - Toplam", 0))
                grup_toplamlari[firma] += top
                c_bf = ws.cell(row=current_row, column=c_idx, value=bf)
                c_bf.border = thin_border; c_bf.number_format = '#,##0.00'; c_bf.alignment = center_align
                
                if bf > 0:
                    if bf == min_fiyat and min_fiyat != max_fiyat: c_bf.fill = min_fill; c_bf.font = min_font
                    elif bf == max_fiyat and min_fiyat != max_fiyat: c_bf.fill = max_fill; c_bf.font = max_font
                
                c_top = ws.cell(row=current_row, column=c_idx+1, value=top)
                c_top.border = thin_border; c_top.number_format = '#,##0.00'; c_top.alignment = center_align
                c_idx += 2
            current_row += 1
            
        ws.row_dimensions[current_row].height = 25
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        alt_hucre = ws.cell(row=current_row, column=1, value=f"{str(grup).upper()} TOPLAMI")
        alt_hucre.font = bold_font; alt_hucre.fill = subtotal_fill; alt_hucre.border = thin_border; alt_hucre.alignment = center_align
        for i in range(2, 7): ws.cell(row=current_row, column=i).border = thin_border
        
        c_idx = 7
        for firma in firmalar_listesi:
            ws.cell(row=current_row, column=c_idx, value="").border = thin_border
            c_top = ws.cell(row=current_row, column=c_idx+1, value=grup_toplamlari[firma])
            c_top.font = bold_font; c_top.fill = subtotal_fill; c_top.border = thin_border; c_top.number_format = '#,##0.00'; c_top.alignment = center_align
            c_idx += 2
        current_row += 1

    current_row += 1
    ws.row_dimensions[current_row].height = 30
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    opt_hucre = ws.cell(row=current_row, column=1, value=f"TÜM KALEMLER İÇİN KARMA (OPTİMUM) MİNİMUM MALİYET :  {optimum_total:,.2f} ₺")
    opt_hucre.font = Font(bold=True, size=12, color="155724"); opt_hucre.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"); opt_hucre.alignment = left_align
    for i in range(1, 9): ws.cell(row=current_row, column=i).border = thin_border

    ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25; ws.column_dimensions['E'].width = 10; ws.column_dimensions['F'].width = 12
    for i in range(7, 7 + (len(firmalar_listesi)*2)): ws.column_dimensions[get_column_letter(i)].width = 16

    wb.save(output)
    return output.getvalue()

# ==========================================
# 4. TOPLU KIYASLAMA MATRİSİ 
# ==========================================
def build_comparison_matrix():
    df_kalemler = st.session_state.get("is_kalemleri_guncel", st.session_state.is_kalemleri)
    if df_kalemler.empty or not st.session_state.firmalar: return pd.DataFrame()
    
    df_main = df_kalemler.copy()
    df_main['Metraj'] = df_main['Metraj'].apply(safe_float)
    
    for firma in st.session_state.firmalar.keys():
        if firma in st.session_state.teklifler:
            df_firma = st.session_state.teklifler[firma]
            if "İmalat Adı" in df_firma.columns and "Birim Fiyat" in df_firma.columns:
                fiyat_mapping = dict(zip(df_firma["İmalat Adı"], df_firma["Birim Fiyat"].apply(safe_float)))
                bf_col = f"{firma} - B.Fiyat"; top_col = f"{firma} - Toplam"
                df_main[bf_col] = df_main["İmalat Adı"].map(fiyat_mapping).fillna(0.0)
                df_main[top_col] = df_main["Metraj"] * df_main[bf_col]
    return df_main

# ==========================================
# 5. TOPSIS MCDM MOTORU
# ==========================================
def calculate_topsis(df_scores, weights):
    matrix = df_scores.iloc[:, 1:].values.astype(float)
    col_sums = np.sqrt((matrix**2).sum(axis=0))
    col_sums[col_sums == 0] = 1e-9 
    norm_matrix = matrix / col_sums
    w_matrix = norm_matrix * weights
    
    ideal_best = np.array([np.min(w_matrix[:,0]), np.max(w_matrix[:,1]), np.max(w_matrix[:,2]), np.max(w_matrix[:,3])])
    ideal_worst = np.array([np.max(w_matrix[:,0]), np.min(w_matrix[:,1]), np.min(w_matrix[:,2]), np.min(w_matrix[:,3])])
    
    s_best = np.sqrt(((w_matrix - ideal_best)**2).sum(axis=1))
    s_worst = np.sqrt(((w_matrix - ideal_worst)**2).sum(axis=1))
    
    denominator = s_best + s_worst
    c_scores = np.where(denominator == 0, 0, s_worst / denominator)
    
    df_result = pd.DataFrame({
        "Firma": df_scores["Firma"],
        "Toplam Fiyat (TL)": df_scores["Fiyat"].apply(lambda x: tr_format(x)),
        "TOPSIS Skoru": c_scores
    })
    df_result = df_result.sort_values(by="TOPSIS Skoru", ascending=False).reset_index(drop=True)
    df_result.insert(0, "Sıra", df_result.index + 1)
    return df_result

# ==========================================
# ARAYÜZ TASARIMI
# ==========================================
st.title("📊 Dinamik Teklif Kıyaslama ve Karar Destek Motoru")

st.sidebar.header("💾 Proje Hafızası (JSON)")
st.sidebar.download_button("📥 Projeyi Bilgisayara Kaydet (.json)", data=export_project_to_json(), file_name="teklif_projesi.json", mime="application/json", use_container_width=True)
st.sidebar.divider()
uploaded_json = st.sidebar.file_uploader("📂 Mevcut Projeyi Yükle (.json)", type=["json"])
if uploaded_json is not None:
    if st.sidebar.button("Yükle ve Verileri Güncelle"):
        if import_project_from_json(uploaded_json): st.sidebar.success("Yüklendi!"); st.rerun()
        else: st.sidebar.error("Bozuk JSON.")

tab1, tab2, tab3, tab4 = st.tabs(["🏗️ Proje & İş Kalemleri", "🏢 Firma Yönetimi", "🧠 MCDM & Optimizasyon", "🖨️ Kıyaslama Raporu (A3)"])

with tab1:
    st.subheader("Proje Bilgileri ve Metrajlar")
    st.session_state.proje_adi = st.text_input("Proje Adı:", value=st.session_state.proje_adi)
    
    col_a, col_b = st.columns([1, 1])
    with col_a:
        df_indirme_icin = st.session_state.get("is_kalemleri_guncel", st.session_state.is_kalemleri)
        st.download_button("📥 Boş Şablonu İndir (E-Tablo / Excel)", data=generate_empty_template(df_indirme_icin, st.session_state.proje_adi), file_name="Bos_Teklif_Sablonu.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_b:
        is_kalemi_sablonu = st.file_uploader("📂 Doldurulmuş Şablonu Sisteme Yükle", type=["xlsx"])
        if is_kalemi_sablonu is not None and st.button("Şablonu Aktar"):
            df_yeni = pd.read_excel(is_kalemi_sablonu)
            if "Metraj" in df_yeni.columns: df_yeni["Metraj"] = df_yeni["Metraj"].apply(safe_float)
            st.session_state.is_kalemleri = df_yeni; st.success("Aktarıldı!"); st.rerun()

    st.markdown("**Manuel Veri Düzenleme:**")
    st.session_state.is_kalemleri_guncel = st.data_editor(
        st.session_state.is_kalemleri, 
        num_rows="dynamic", 
        use_container_width=True,
        key="editor_kalemler_ui"
    )

with tab2:
    st.subheader("Firma Bilgileri ve Teklif Yükleme")
    with st.form("firma_ekle_form"):
        f_ad = st.text_input("Firma Adı:")
        col_f1, col_f2 = st.columns(2)
        f_yetkili = col_f1.text_input("Yetkili Kişi:"); f_tel = col_f2.text_input("Telefon:")
        col_f3, col_f4 = st.columns(2)
        f_tarih = col_f3.date_input("Teklif Tarihi:")
        f_gecerlilik = col_f4.date_input("Geçerlilik Tarihi:")
        f_not = st.text_area("Firma Hakkında Kısa Not / Gözlem:")
        if st.form_submit_button("➕ Firmayı Kaydet"):
            if f_ad and f_ad not in st.session_state.firmalar:
                st.session_state.firmalar[f_ad.strip()] = {
                    "yetkili": f_yetkili, "telefon": f_tel, "not": f_not, 
                    "teklif_tarihi": f_tarih.strftime("%d.%m.%Y"), "gecerlilik_tarihi": f_gecerlilik.strftime("%d.%m.%Y")
                }
                st.success("Eklendi."); st.rerun()
                
    st.divider()
    if st.session_state.firmalar:
        secili_firma = st.selectbox("İşlem Yapılacak / Teklifi Yüklenecek Firma:", list(st.session_state.firmalar.keys()))
        
        f_bilgi = st.session_state.firmalar[secili_firma]
        st.markdown(f"""
        **Yetkili:** {f_bilgi.get('yetkili', '-')} | **Telefon:** {f_bilgi.get('telefon', '-')}  
        **Teklif Tarihi:** {f_bilgi.get('teklif_tarihi', '-')} | **Geçerlilik:** {f_bilgi.get('gecerlilik_tarihi', '-')}  
        *Not:* {f_bilgi.get('not', '-')}
        """)
        
        yuklenen_dosya = st.file_uploader(f"{secili_firma} - İmzalı Teklif Dosyası Yükle", type=["xlsx"])
        if yuklenen_dosya is not None and st.button("Teklifi Sisteme Kaydet"):
            st.session_state.teklifler[secili_firma] = pd.read_excel(yuklenen_dosya)
            st.success("Teklif alındı!")
            
        if secili_firma in st.session_state.teklifler:
            st.success(f"✅ {secili_firma} firmasına ait teklif sisteme yüklenmiştir. (Aşağıdan inceleyebilirsiniz)")
            df_firma_teklif = st.session_state.teklifler[secili_firma]
            
            col_tablo, col_indir = st.columns([4, 1])
            with col_tablo:
                st.dataframe(df_firma_teklif, use_container_width=True, height=250)
            with col_indir:
                st.download_button(
                    label=f"📥 {secili_firma} Teklifini İndir", 
                    data=generate_firm_excel(df_firma_teklif), 
                    file_name=f"{secili_firma}_Teklifi.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    use_container_width=True
                )

with tab3:
    st.header("🧠 Karar Destek: Karma Optimizasyon ve TOPSIS MCDM")
    df_comp = build_comparison_matrix()
    
    if not df_comp.empty and st.session_state.teklifler:
        top_cols = [c for c in df_comp.columns if " - Toplam" in c]
        bf_cols = [c for c in df_comp.columns if " - B.Fiyat" in c]
        firm_totals = df_comp[top_cols].sum().rename(index=lambda x: x.replace(" - Toplam", ""))
        firm_totals = firm_totals[firm_totals > 0].sort_values()
        
        st.subheader("🏆 Firma Genel Sıralaması")
        if not firm_totals.empty:
            cols = st.columns(len(firm_totals))
            for i, (firma, total) in enumerate(firm_totals.items()):
                with cols[i]:
                    if i == 0: st.success(f"🥇 1. {firma}\n\n**{tr_format(total)} ₺**")
                    elif i == 1: st.warning(f"🥈 2. {firma}\n\n**{tr_format(total)} ₺**")
                    elif i == 2: st.info(f"🥉 3. {firma}\n\n**{tr_format(total)} ₺**")
                    else: st.markdown(f"**{i+1}. {firma}**\n\n{tr_format(total)} ₺")
        st.divider()

        st.subheader("🎨 Detaylı Kıyaslama ve Karma Optimizasyon")
        if bf_cols:
            df_opt = df_comp[["İmalat Adı", "Metraj"] + bf_cols].copy()
            df_opt_zero_nan = df_opt[bf_cols].replace(0, np.nan)
            df_opt["En Düşük B.Fiyat"] = df_opt_zero_nan.min(axis=1).fillna(0)
            df_opt["Optimum Toplam"] = df_opt["Metraj"] * df_opt["En Düşük B.Fiyat"]
            
            def highlight_min_max(row):
                styles = [''] * len(row)
                numeric_row = pd.to_numeric(row, errors='coerce')
                min_val = numeric_row[numeric_row > 0].min(); max_val = numeric_row.max()
                for i, val in enumerate(numeric_row):
                    if pd.notna(val) and val > 0:
                        if val == min_val and min_val != max_val: styles[i] = 'background-color: #d4edda; color: #155724;'
                        elif val == max_val and min_val != max_val: styles[i] = 'background-color: #f8d7da; color: #721c24;'
                return styles
            
            format_dict = {"Metraj": "{:,.2f}", "En Düşük B.Fiyat": lambda x: f"{tr_format(x)} ₺", "Optimum Toplam": lambda x: f"{tr_format(x)} ₺"}
            for c in bf_cols: format_dict[c] = lambda x: f"{tr_format(x)} ₺"
                
            styled_df = df_opt.style.apply(highlight_min_max, subset=bf_cols, axis=1).format(format_dict)
            st.dataframe(styled_df, use_container_width=True)
            
            st.success(f"✨ **Karma Strateji (Optimum Dağılım):** Tüm işleri kalem bazında en ucuz fiyatı veren firmalarla çalışırsanız Toplam Maliyetiniz **{tr_format(df_opt['Optimum Toplam'].sum())} ₺** olacaktır.")
        st.divider()

        st.subheader("⚖️ Çok Kriterli Karar Verme (TOPSIS)")
        col_w1, col_w2, col_w3, col_w4 = st.columns(4)
        
        w_fiyat = col_w1.number_input("Fiyat Ağırlığı (%)", min_value=0, max_value=100, value=50)
        w_kalite = col_w2.number_input("Kalite/Teknik (%)", min_value=0, max_value=100, value=20)
        w_finans = col_w3.number_input("Finansal Güç (%)", min_value=0, max_value=100, value=17)
        w_hiz = col_w4.number_input("Teslimat Hızı (%)", min_value=0, max_value=100, value=13)
        
        toplam_agirlik = w_fiyat + w_kalite + w_finans + w_hiz
        if toplam_agirlik != 100:
            st.error(f"🛑 **HATA:** Kriter ağırlıklarının toplamı tam **%100** olmalıdır. (Şu anki toplam: **%{toplam_agirlik}**) Lütfen yukarıdaki kutulardan değerleri düzeltin.")
            topsis_hazir = False
        else:
            st.success("✅ Ağırlıklar dengeli (%100). Analize hazır.")
            topsis_hazir = True
            
        st.markdown("**Firma Değerlendirme Puanları (1 En Kötü - 10 En İyi)**")
        
        current_firms = list(firm_totals.index)
        if "mcdm_base_df" not in st.session_state or list(st.session_state.mcdm_base_df.get("Firma", [])) != current_firms:
            st.session_state.mcdm_base_df = pd.DataFrame({
                "Firma": current_firms,
                "Fiyat (Bilgi Amaçlı)": [tr_format(firm_totals[f]) for f in current_firms],
                "Kalite Puanı (1-10)": 5, "Finansal Puan (1-10)": 5, "Hız Puanı (1-10)": 5
            })
            
        edited_mcdm = st.data_editor(
            st.session_state.mcdm_base_df, 
            hide_index=True, 
            disabled=["Firma", "Fiyat (Bilgi Amaçlı)"], 
            use_container_width=True,
            key="editor_mcdm_ui" 
        )
        
        if st.button("🚀 TOPSIS Analizini Çalıştır", type="primary", disabled=not topsis_hazir):
            math_df = pd.DataFrame({
                "Firma": edited_mcdm["Firma"],
                "Fiyat": [firm_totals[f] for f in edited_mcdm["Firma"]],
                "Kalite": edited_mcdm["Kalite Puanı (1-10)"],
                "Finans": edited_mcdm["Finansal Puan (1-10)"],
                "Hiz": edited_mcdm["Hız Puanı (1-10)"]
            })
            weights = np.array([w_fiyat/100, w_kalite/100, w_finans/100, w_hiz/100])
            sonuc_df = calculate_topsis(math_df, weights)
            st.success("✅ Analiz Tamamlandı!")
            
            c_tab, c_chart = st.columns([1, 1])
            with c_tab:
                st.dataframe(sonuc_df.style.format({"TOPSIS Skoru": "{:.4f}"}), hide_index=True, use_container_width=True)
                st.markdown(f"🏆 **Optimal Tercih:** **{sonuc_df.iloc[0]['Firma']}**")
            
            with c_chart:
                fig_radar = go.Figure()
                min_p = math_df["Fiyat"].min(); max_p = math_df["Fiyat"].max()
                colors = ['rgba(44, 123, 229, 0.15)', 'rgba(0, 217, 126, 0.15)', 'rgba(230, 55, 87, 0.15)', 'rgba(245, 128, 62, 0.15)']
                line_colors = ['rgb(44, 123, 229)', 'rgb(0, 217, 126)', 'rgb(230, 55, 87)', 'rgb(245, 128, 62)']
                
                for i, row in math_df.iterrows():
                    p_score = 10 if max_p == min_p else 10 - 9 * ((row["Fiyat"] - min_p) / (max_p - min_p))
                    r_vals = [p_score, row["Kalite"], row["Finans"], row["Hiz"]]
                    r_vals.append(r_vals[0])
                    theta_vals = ['Fiyat Puanı', 'Kalite Puanı', 'Finansal Puan', 'Hız Puanı', 'Fiyat Puanı']
                    
                    fig_radar.add_trace(go.Scatterpolar(
                        r=r_vals, theta=theta_vals, fill='toself', name=row["Firma"],
                        fillcolor=colors[i % len(colors)], line=dict(color=line_colors[i % len(line_colors)], width=2),
                        hoverinfo="text", text=[f"{t}: {v:.1f}" for t, v in zip(theta_vals, r_vals)],
                        marker=dict(size=6, color=line_colors[i % len(line_colors)])
                    ))
                    
                fig_radar.update_layout(
                    polar=dict(
                        radialaxis=dict(visible=True, range=[0, 10], showticklabels=False, gridcolor='rgba(0,0,0,0.05)', linecolor='rgba(0,0,0,0.05)'),
                        angularaxis=dict(gridcolor='rgba(0,0,0,0.05)', linecolor='rgba(0,0,0,0.05)', tickfont=dict(size=12, color='gray'))
                    ),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5, font=dict(color='gray')),
                    margin=dict(l=40, r=40, t=20, b=20)
                )
                st.plotly_chart(fig_radar, use_container_width=True)
    else:
        st.warning("Analiz için teklif verisi gereklidir.")

with tab4:
    st.subheader("📄 Profesyonel Kıyaslama Raporu (A3 Formatı)")
    df_comp = build_comparison_matrix()
    if not df_comp.empty and st.session_state.teklifler:
        col_btn1, col_btn2 = st.columns(2)
        
        df_indirme_icin = st.session_state.get("is_kalemleri_guncel", st.session_state.is_kalemleri)
        
        excel_data_orijinal = generate_a3_excel_report(df_comp, df_indirme_icin, st.session_state.firmalar, anonim=False)
        col_btn1.download_button("📄 Orijinal Raporu İndir (A3 Excel)", data=excel_data_orijinal, file_name="Orijinal_Kiyaslama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
        
        excel_data_anonim = generate_a3_excel_report(df_comp, df_indirme_icin, st.session_state.firmalar, anonim=True)
        col_btn2.download_button("🕵️ Anonim Raporu İndir (Firma Adları Gizli)", data=excel_data_anonim, file_name="Anonim_Kiyaslama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        st.info("Kıyaslama tablosunun oluşması için iş kalemi ve teklif yüklemelisiniz.")
